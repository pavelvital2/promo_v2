from __future__ import annotations

from datetime import UTC, datetime, timedelta
import shutil
import tempfile

from django.apps import apps
from django.contrib.auth import get_user_model
from django.core.exceptions import PermissionDenied, ValidationError
from django.core.files.storage import default_storage
from django.test import TestCase, override_settings
from openpyxl import load_workbook

from apps.audit.models import AuditRecord
from apps.discounts.wb_api.redaction import contains_secret_like
from apps.files.models import FileObject
from apps.identity_access.models import AccessEffect, Permission, Role, StoreAccess, UserPermissionOverride
from apps.identity_access.seeds import ROLE_MARKETPLACE_MANAGER, seed_identity_access
from apps.operations.models import (
    Marketplace,
    Operation,
    OperationDetailRow,
    OperationMode,
    OperationOutputFile,
    OperationStepCode,
    OperationType,
    OutputKind,
)
from apps.operations.services import create_run
from apps.stores.models import ConnectionBlock, StoreAccount
from apps.stores.services import WB_API_CONNECTION_TYPE, WB_API_MODULE
from apps.techlog.models import TechLogRecord

from .models import WBPromotion, WBPromotionExportFile, WBPromotionProduct, WBPromotionSnapshot
from .normalizers import REASON_PRODUCT_INVALID, REASON_REGULAR, normalize_promotion
from .services import DETAILS_BATCH_SIZE, NOMENCLATURES_LIMIT, PROMOTIONS_LIMIT, download_wb_current_promotions


class FakePromotionsClient:
    def __init__(self, *, pages, details=None, nomenclatures=None):
        self.pages = list(pages)
        self.details = details or {}
        self.nomenclatures = {
            key: list(value)
            for key, value in (nomenclatures or {}).items()
        }
        self.calls = []

    def list_promotions(self, *, start_datetime, end_datetime, all_promo, limit, offset):
        self.calls.append(
            {
                "endpoint": "list",
                "startDateTime": start_datetime,
                "endDateTime": end_datetime,
                "allPromo": all_promo,
                "limit": limit,
                "offset": offset,
            },
        )
        return {"data": {"promotions": self.pages.pop(0)}}

    def promotion_details(self, *, promotion_ids):
        self.calls.append({"endpoint": "details", "promotionIDs": list(promotion_ids)})
        return {
            "data": {
                "promotions": [
                    self.details.get(promotion_id, {"id": promotion_id, "description": "safe"})
                    for promotion_id in promotion_ids
                ],
            },
        }

    def promotion_nomenclatures(self, *, promotion_id, in_action, limit, offset):
        self.calls.append(
            {
                "endpoint": "nomenclatures",
                "promotionID": promotion_id,
                "inAction": in_action,
                "limit": limit,
                "offset": offset,
            },
        )
        return {"data": {"nomenclatures": self.nomenclatures[(promotion_id, in_action)].pop(0)}}


class FakeClientFactory:
    def __init__(self, *, pages, details=None, nomenclatures=None):
        self.client = FakePromotionsClient(
            pages=pages,
            details=details,
            nomenclatures=nomenclatures,
        )
        self.kwargs = None

    def __call__(self, **kwargs):
        self.kwargs = kwargs
        return self.client


class WBApiPromotionsTask013Tests(TestCase):
    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls._media_root = tempfile.mkdtemp(prefix="promo-v2-wb-api-promotions-")
        cls.override = override_settings(MEDIA_ROOT=cls._media_root)
        cls.override.enable()

    @classmethod
    def tearDownClass(cls):
        cls.override.disable()
        shutil.rmtree(cls._media_root, ignore_errors=True)
        super().tearDownClass()

    def setUp(self):
        seed_identity_access()
        role = Role.objects.get(code=ROLE_MARKETPLACE_MANAGER)
        User = get_user_model()
        self.user = User.objects.create_user(
            login=f"manager-promos-{self._testMethodName}",
            password="password",
            display_name="Manager",
            primary_role=role,
        )
        self.store = StoreAccount.objects.create(
            name=f"WB Promotions Store {self._testMethodName}",
            marketplace=StoreAccount.Marketplace.WB,
        )
        StoreAccess.objects.create(
            user=self.user,
            store=self.store,
            access_level=StoreAccess.AccessLevel.WORK,
            effect=AccessEffect.ALLOW,
        )
        self.connection = ConnectionBlock.objects.create(
            store=self.store,
            module=WB_API_MODULE,
            connection_type=WB_API_CONNECTION_TYPE,
            status=ConnectionBlock.Status.ACTIVE,
            protected_secret_ref="env://TASK013_WB_TOKEN",
            metadata={"label": "safe"},
            is_stage2_1_used=True,
        )
        self.now = datetime(2026, 4, 26, 9, 0, tzinfo=UTC)

    @staticmethod
    def _secret_resolver(protected_secret_ref):
        assert protected_secret_ref == "env://TASK013_WB_TOKEN"
        return "Bearer task013-local-token-value"

    def _promotion(self, promotion_id, *, start_delta_hours=-1, end_delta_hours=1, promotion_type="regular"):
        return {
            "id": promotion_id,
            "name": f"Promo {promotion_id}",
            "type": promotion_type,
            "startDateTime": (self.now + timedelta(hours=start_delta_hours)).isoformat().replace("+00:00", "Z"),
            "endDateTime": (self.now + timedelta(hours=end_delta_hours)).isoformat().replace("+00:00", "Z"),
        }

    def _product(self, nm_id=501, *, plan_price=900, plan_discount=10, in_action=True):
        return {
            "id": nm_id,
            "inAction": in_action,
            "price": 1000,
            "currencyCode": "RUB",
            "planPrice": plan_price,
            "discount": 5,
            "planDiscount": plan_discount,
        }

    def test_current_filter_boundaries_are_start_inclusive_end_exclusive(self):
        start_equal = normalize_promotion(
            self._promotion(1, start_delta_hours=0, end_delta_hours=1),
            now_utc=self.now,
        )
        end_equal = normalize_promotion(
            self._promotion(2, start_delta_hours=-1, end_delta_hours=0),
            now_utc=self.now,
        )

        self.assertTrue(start_equal.is_current_at_fetch)
        self.assertFalse(end_equal.is_current_at_fetch)

    def test_download_uses_all_promo_window_saves_timestamps_and_exports_excel(self):
        factory = FakeClientFactory(
            pages=[[self._promotion(101), self._promotion(102, end_delta_hours=0)], []],
            details={101: {"id": 101, "description": "regular-safe"}},
            nomenclatures={
                (101, True): [[self._product(501, in_action=True)], []],
                (101, False): [[self._product(502, in_action=False, plan_price=None)], []],
            },
        )

        operation = download_wb_current_promotions(
            actor=self.user,
            store=self.store,
            client_factory=factory,
            secret_resolver=self._secret_resolver,
            now_utc=self.now,
        )

        list_calls = [call for call in factory.client.calls if call["endpoint"] == "list"]
        self.assertEqual(list_calls[0]["allPromo"], True)
        self.assertEqual(list_calls[0]["limit"], PROMOTIONS_LIMIT)
        self.assertIn("2026-04-25T09:00:00Z", list_calls[0]["startDateTime"])
        self.assertIn("2026-04-27T09:00:00Z", list_calls[0]["endDateTime"])
        self.assertEqual(operation.summary["current_filter_timestamp"], self.now.isoformat())
        self.assertEqual(operation.summary["allPromo"], True)
        self.assertEqual(operation.summary["current_promotions_count"], 1)
        self.assertEqual(operation.step_code, OperationStepCode.WB_API_PROMOTIONS_DOWNLOAD)
        self.assertEqual(operation.operation_type, OperationType.NOT_APPLICABLE)

        snapshot = WBPromotionSnapshot.objects.get(operation=operation)
        self.assertEqual(snapshot.store, self.store)
        self.assertEqual(snapshot.current_filter_timestamp, self.now)
        self.assertEqual(snapshot.api_window_start, self.now - timedelta(hours=24))
        self.assertEqual(snapshot.api_window_end, self.now + timedelta(hours=24))
        self.assertEqual(snapshot.promotions_count, 2)
        self.assertEqual(snapshot.current_promotions_count, 1)
        self.assertEqual(snapshot.regular_current_promotions_count, 1)
        self.assertEqual(snapshot.auto_current_promotions_count, 0)
        self.assertEqual(snapshot.promotion_products_count, 2)
        self.assertEqual(snapshot.invalid_product_count, 1)
        self.assertEqual(operation.summary["wb_promotion_snapshot_id"], snapshot.pk)
        self.assertIn("101", snapshot.raw_response_safe_snapshot["current_details"])

        current_promotion = WBPromotion.objects.get(store=self.store, wb_promotion_id=101)
        filtered_promotion = WBPromotion.objects.get(store=self.store, wb_promotion_id=102)
        self.assertEqual(current_promotion.snapshot_ref, snapshot)
        self.assertTrue(current_promotion.is_current_at_fetch)
        self.assertFalse(filtered_promotion.is_current_at_fetch)
        self.assertEqual(current_promotion.type, "regular")

        output_links = OperationOutputFile.objects.filter(operation=operation)
        self.assertEqual(output_links.count(), 1)
        output = output_links.get().file_version
        self.assertEqual(output_links.get().output_kind, OutputKind.PROMOTION_EXPORT)
        self.assertEqual(output.file.scenario, FileObject.Scenario.WB_DISCOUNTS_API_PROMOTION_EXPORT)
        with default_storage.open(output.storage_path, "rb") as stream:
            workbook = load_workbook(stream)
            sheet = workbook.active
            self.assertEqual(
                [cell.value for cell in sheet[1]],
                [
                    "Артикул WB",
                    "Плановая цена для акции",
                    "Загружаемая скидка для участия в акции",
                ],
            )
            self.assertEqual(sheet["A2"].value, "501")
            self.assertEqual(sheet["B2"].value, "900")
            self.assertEqual(sheet["C2"].value, 10)
            self.assertEqual(sheet["A3"].value, "502")
            self.assertEqual(sheet["B3"].value, None)
            self.assertIn("_api_raw", workbook.sheetnames)
            workbook.close()

        invalid_detail = OperationDetailRow.objects.get(
            operation=operation,
            product_ref="502",
            reason_code=REASON_PRODUCT_INVALID,
        )
        self.assertEqual(invalid_detail.problem_field, "planPrice/planDiscount")

        persisted_products = WBPromotionProduct.objects.filter(
            source_snapshot=snapshot,
            promotion=current_promotion,
        ).order_by("nmID")
        self.assertEqual(persisted_products.count(), 2)
        self.assertEqual(persisted_products[0].nmID, "501")
        self.assertEqual(persisted_products[0].row_status, "valid")
        self.assertEqual(persisted_products[1].nmID, "502")
        self.assertEqual(persisted_products[1].reason_code, REASON_PRODUCT_INVALID)

        export_link = WBPromotionExportFile.objects.get(operation=operation)
        self.assertEqual(export_link.promotion, current_promotion)
        self.assertEqual(export_link.file_version, output)

    def test_details_are_batched_by_100_unique_ids_and_auto_promotions_have_no_products(self):
        promotions = [self._promotion(promotion_id, promotion_type="auto") for promotion_id in range(1, 103)]
        promotions.append(self._promotion(102, promotion_type="auto"))
        factory = FakeClientFactory(pages=[promotions, []])

        operation = download_wb_current_promotions(
            actor=self.user,
            store=self.store,
            client_factory=factory,
            secret_resolver=self._secret_resolver,
            now_utc=self.now,
        )

        detail_calls = [call for call in factory.client.calls if call["endpoint"] == "details"]
        self.assertEqual([len(call["promotionIDs"]) for call in detail_calls], [DETAILS_BATCH_SIZE, 2])
        self.assertFalse(any(call["endpoint"] == "nomenclatures" for call in factory.client.calls))
        self.assertEqual(OperationOutputFile.objects.filter(operation=operation).count(), 0)
        self.assertEqual(operation.summary["auto_current_promotions_count"], 102)
        snapshot = WBPromotionSnapshot.objects.get(operation=operation)
        self.assertEqual(snapshot.current_promotions_count, 102)
        self.assertEqual(snapshot.auto_current_promotions_count, 102)
        self.assertEqual(snapshot.promotion_products_count, 0)
        self.assertEqual(WBPromotion.objects.filter(snapshot_ref=snapshot).count(), 102)
        self.assertFalse(WBPromotionProduct.objects.filter(source_snapshot=snapshot).exists())
        self.assertFalse(WBPromotionExportFile.objects.filter(operation=operation).exists())
        self.assertFalse(
            OperationDetailRow.objects.filter(
                operation=operation,
                reason_code__in=[
                    "wb_api_promotion_product_valid",
                    "wb_api_promotion_product_invalid",
                ],
            ).exists(),
        )

    def test_regular_promotions_without_products_do_not_create_empty_excel_files(self):
        factory = FakeClientFactory(
            pages=[[self._promotion(301), self._promotion(302)], []],
            nomenclatures={
                (301, True): [[],],
                (301, False): [[],],
                (302, True): [[self._product(701)], []],
                (302, False): [[],],
            },
        )

        operation = download_wb_current_promotions(
            actor=self.user,
            store=self.store,
            client_factory=factory,
            secret_resolver=self._secret_resolver,
            now_utc=self.now,
        )

        snapshot = WBPromotionSnapshot.objects.get(operation=operation)
        self.assertEqual(snapshot.regular_current_promotions_count, 2)
        self.assertEqual(snapshot.promotion_products_count, 1)
        self.assertEqual(OperationOutputFile.objects.filter(operation=operation).count(), 1)
        self.assertEqual(WBPromotionExportFile.objects.filter(operation=operation).count(), 1)
        self.assertEqual(operation.summary["output_file_version_ids"], [OperationOutputFile.objects.get(operation=operation).file_version_id])

        empty_detail = OperationDetailRow.objects.get(
            operation=operation,
            product_ref="301",
            reason_code=REASON_REGULAR,
        )
        self.assertEqual(empty_detail.final_value["products_count"], 0)
        self.assertIsNone(empty_detail.final_value["output_file_version_id"])
        self.assertIn("no empty Excel export", empty_detail.message)

    def test_nomenclatures_are_requested_for_true_and_false_until_empty_page(self):
        factory = FakeClientFactory(
            pages=[[self._promotion(201)], []],
            nomenclatures={
                (201, True): [[self._product(601)], [self._product(602)], []],
                (201, False): [[self._product(603, in_action=False)], []],
            },
        )

        download_wb_current_promotions(
            actor=self.user,
            store=self.store,
            client_factory=factory,
            secret_resolver=self._secret_resolver,
            now_utc=self.now,
        )

        calls = [call for call in factory.client.calls if call["endpoint"] == "nomenclatures"]
        self.assertEqual(
            [(call["inAction"], call["limit"], call["offset"]) for call in calls],
            [
                (True, NOMENCLATURES_LIMIT, 0),
                (True, NOMENCLATURES_LIMIT, NOMENCLATURES_LIMIT),
                (True, NOMENCLATURES_LIMIT, NOMENCLATURES_LIMIT * 2),
                (False, NOMENCLATURES_LIMIT, 0),
                (False, NOMENCLATURES_LIMIT, NOMENCLATURES_LIMIT),
            ],
        )

    def test_permission_object_access_and_active_connection_required(self):
        User = get_user_model()
        outsider = User.objects.create_user(
            login=f"outsider-promos-{self._testMethodName}",
            password="password",
            display_name="Outsider",
            primary_role=Role.objects.get(code=ROLE_MARKETPLACE_MANAGER),
        )
        with self.assertRaises(PermissionDenied):
            download_wb_current_promotions(
                actor=outsider,
                store=self.store,
                client_factory=FakeClientFactory(pages=[[]]),
                secret_resolver=self._secret_resolver,
                now_utc=self.now,
            )

        self.connection.status = ConnectionBlock.Status.CONFIGURED
        self.connection.save(update_fields=["status", "updated_at"])
        with self.assertRaises(PermissionDenied):
            download_wb_current_promotions(
                actor=self.user,
                store=self.store,
                client_factory=FakeClientFactory(pages=[[]]),
                secret_resolver=self._secret_resolver,
                now_utc=self.now,
            )

        self.connection.status = ConnectionBlock.Status.ACTIVE
        self.connection.save(update_fields=["status", "updated_at"])
        permission = Permission.objects.get(code="wb.api.promotions.download")
        UserPermissionOverride.objects.create(
            user=self.user,
            store=self.store,
            permission=permission,
            effect=AccessEffect.DENY,
        )
        with self.assertRaises(PermissionDenied):
            download_wb_current_promotions(
                actor=self.user,
                store=self.store,
                client_factory=FakeClientFactory(pages=[[]]),
                secret_resolver=self._secret_resolver,
                now_utc=self.now,
            )

    def test_secret_redaction_in_operation_audit_techlog_snapshots_and_files(self):
        operation = download_wb_current_promotions(
            actor=self.user,
            store=self.store,
            client_factory=FakeClientFactory(
                pages=[[self._promotion(301)], []],
                nomenclatures={(301, True): [[self._product(701)], []], (301, False): [[]]},
            ),
            secret_resolver=self._secret_resolver,
            now_utc=self.now,
        )
        combined = str(
            {
                "operation": Operation.objects.filter(pk=operation.pk).values(
                    "execution_context",
                    "summary",
                ).get(),
                "audit": list(
                    AuditRecord.objects.filter(operation=operation).values(
                        "safe_message",
                        "before_snapshot",
                        "after_snapshot",
                    ),
                ),
                "techlog": list(
                    TechLogRecord.objects.filter(operation=operation).values(
                        "safe_message",
                        "sensitive_details_ref",
                    ),
                ),
                "files": list(
                    OperationOutputFile.objects.filter(operation=operation).values(
                        "file_version__original_name",
                        "file_version__file__logical_name",
                    ),
                ),
            },
        )

        self.assertFalse(contains_secret_like(combined))
        self.assertNotIn("task013-local-token-value", combined)
        self.assertNotIn("Authorization", combined)

    def test_operation_classifier_rejects_check_process_for_promotions_api_step(self):
        run = create_run(
            marketplace=Marketplace.WB,
            module="wb_api",
            mode=OperationMode.API,
            store=self.store,
            initiated_by=self.user,
        )
        operation = Operation(
            marketplace=Marketplace.WB,
            module="wb_api",
            mode=OperationMode.API,
            operation_type=OperationType.PROCESS,
            step_code=OperationStepCode.WB_API_PROMOTIONS_DOWNLOAD,
            status="created",
            run=run,
            store=self.store,
            initiator_user=self.user,
            logic_version="test",
        )

        with self.assertRaises(ValidationError):
            operation.full_clean()

    def test_dedicated_promotion_models_are_discoverable(self):
        self.assertIs(apps.get_model("promotions", "WBPromotion"), WBPromotion)
        self.assertIs(apps.get_model("promotions", "WBPromotionSnapshot"), WBPromotionSnapshot)
        self.assertIs(apps.get_model("promotions", "WBPromotionProduct"), WBPromotionProduct)
        self.assertIs(apps.get_model("promotions", "WBPromotionExportFile"), WBPromotionExportFile)

    def test_no_wb_promotions_upload_endpoint_is_referenced(self):
        from pathlib import Path

        root = Path(__file__).resolve().parents[4]
        forbidden = "/api/v1/calendar/promotions/" + "upload"
        code = "\n".join(
            path.read_text(encoding="utf-8")
            for path in (root / "apps").rglob("*.py")
            if "__pycache__" not in str(path)
        )
        self.assertNotIn(forbidden.lstrip("/"), code)
        self.assertNotIn(forbidden, code)
