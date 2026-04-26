from __future__ import annotations

import shutil
import tempfile

from django.contrib.auth import get_user_model
from django.core.exceptions import PermissionDenied, ValidationError
from django.core.files.storage import default_storage
from django.test import TestCase, override_settings
from openpyxl import load_workbook

from apps.audit.models import AuditRecord
from apps.discounts.wb_api.redaction import contains_secret_like
from apps.files.models import FileObject
from apps.identity_access.models import AccessEffect, Role, StoreAccess, UserPermissionOverride
from apps.identity_access.models import Permission
from apps.identity_access.seeds import ROLE_MARKETPLACE_MANAGER, seed_identity_access
from apps.marketplace_products.models import MarketplaceProduct, MarketplaceProductHistory
from apps.operations.models import (
    Marketplace,
    Operation,
    OperationDetailRow,
    OperationMode,
    OperationOutputFile,
    OperationStepCode,
    OperationType,
)
from apps.operations.services import create_run
from apps.stores.models import ConnectionBlock, StoreAccount
from apps.stores.services import WB_API_CONNECTION_TYPE, WB_API_MODULE
from apps.techlog.models import TechLogRecord

from .normalizers import REASON_INVALID, REASON_SIZE_CONFLICT, REASON_VALID, normalize_price_good
from .services import PRICES_LIMIT, download_wb_prices


class FakePricesClient:
    def __init__(self, *, token, store_scope, pages):
        self.token = token
        self.store_scope = store_scope
        self.pages = list(pages)
        self.calls = []

    def list_goods_filter(self, *, limit, offset):
        self.calls.append({"limit": limit, "offset": offset})
        return {"data": {"listGoods": self.pages.pop(0)}}


class FakeClientFactory:
    def __init__(self, pages):
        self.client = None
        self.pages = pages

    def __call__(self, *, token, store_scope):
        self.client = FakePricesClient(token=token, store_scope=store_scope, pages=self.pages)
        return self.client


class WBApiPricesTask012Tests(TestCase):
    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls._media_root = tempfile.mkdtemp(prefix="promo-v2-wb-api-prices-")
        cls.override = override_settings(MEDIA_ROOT=cls._media_root)
        cls.override.enable()

    @classmethod
    def tearDownClass(cls):
        cls.override.disable()
        shutil.rmtree(cls._media_root, ignore_errors=True)
        super().tearDownClass()

    def setUp(self):
        seed_identity_access()
        role = Role.objects.get(code=ROLE_MARKETPLACE_MANAGER)
        User = get_user_model()
        self.user = User.objects.create_user(
            login=f"manager-{self._testMethodName}",
            password="password",
            display_name="Manager",
            primary_role=role,
        )
        self.store = StoreAccount.objects.create(
            name=f"WB API Store {self._testMethodName}",
            marketplace=StoreAccount.Marketplace.WB,
        )
        StoreAccess.objects.create(
            user=self.user,
            store=self.store,
            access_level=StoreAccess.AccessLevel.WORK,
            effect=AccessEffect.ALLOW,
        )
        self.connection = ConnectionBlock.objects.create(
            store=self.store,
            module=WB_API_MODULE,
            connection_type=WB_API_CONNECTION_TYPE,
            status=ConnectionBlock.Status.ACTIVE,
            protected_secret_ref="env://TASK012_WB_TOKEN",
            metadata={"label": "safe"},
            is_stage2_1_used=True,
        )

    @staticmethod
    def _secret_resolver(protected_secret_ref):
        assert protected_secret_ref == "env://TASK012_WB_TOKEN"
        return "Bearer task012-local-token-value"

    def _good(self, nm_id=101, price=1000, *, second_price=None, sizes=True):
        size_rows = [] if not sizes else [
            {
                "sizeID": 1,
                "price": price,
                "discountedPrice": 900,
                "clubDiscountedPrice": 850,
                "techSizeName": "0",
            },
            {
                "sizeID": 2,
                "price": price if second_price is None else second_price,
                "discountedPrice": 900,
                "clubDiscountedPrice": 850,
                "techSizeName": "1",
            },
        ]
        return {
            "nmID": nm_id,
            "vendorCode": f"vendor-{nm_id}",
            "sizes": size_rows,
            "currencyIsoCode4217": "RUB",
            "discount": 10,
            "clubDiscount": 15,
            "editableSizePrice": False,
            "isBadTurnover": False,
        }

    def test_normalizer_size_rules_equal_conflict_and_invalid(self):
        valid = normalize_price_good(self._good(price=1000), row_no=1)
        conflict = normalize_price_good(self._good(price=1000, second_price=1100), row_no=2)
        invalid = normalize_price_good(self._good(sizes=False), row_no=3)

        self.assertEqual(valid.reason_code, REASON_VALID)
        self.assertEqual(str(valid.derived_price), "1000")
        self.assertTrue(valid.upload_ready)
        self.assertEqual(conflict.reason_code, REASON_SIZE_CONFLICT)
        self.assertFalse(conflict.upload_ready)
        self.assertIsNone(conflict.derived_price)
        self.assertEqual(invalid.reason_code, REASON_INVALID)
        self.assertFalse(invalid.upload_ready)

    def test_download_paginates_to_empty_page_exports_excel_and_updates_products(self):
        pages = [[self._good(101), self._good(102, price=2000, second_price=2100)], []]
        factory = FakeClientFactory(pages)

        operation = download_wb_prices(
            actor=self.user,
            store=self.store,
            client_factory=factory,
            secret_resolver=self._secret_resolver,
        )

        self.assertEqual(factory.client.calls, [
            {"limit": PRICES_LIMIT, "offset": 0},
            {"limit": PRICES_LIMIT, "offset": PRICES_LIMIT},
        ])
        self.assertEqual(operation.marketplace, Marketplace.WB)
        self.assertEqual(operation.mode, OperationMode.API)
        self.assertEqual(operation.step_code, OperationStepCode.WB_API_PRICES_DOWNLOAD)
        self.assertEqual(operation.operation_type, OperationType.NOT_APPLICABLE)
        self.assertEqual(operation.status, "completed_with_warnings")
        self.assertEqual(OperationOutputFile.objects.filter(operation=operation).count(), 1)

        output = OperationOutputFile.objects.get(operation=operation).file_version
        self.assertEqual(output.file.scenario, FileObject.Scenario.WB_DISCOUNTS_API_PRICE_EXPORT)
        with default_storage.open(output.storage_path, "rb") as stream:
            workbook = load_workbook(stream)
            sheet = workbook.active
            self.assertEqual([cell.value for cell in sheet[1]], ["Артикул WB", "Текущая цена", "Новая скидка"])
            self.assertEqual(sheet["A2"].value, "101")
            self.assertEqual(sheet["B2"].value, "1000")
            self.assertEqual(sheet["C2"].value, None)
            self.assertEqual(sheet["A3"].value, "102")
            self.assertEqual(sheet["B3"].value, None)
            self.assertIn("_api_raw", workbook.sheetnames)
            workbook.close()

        self.assertEqual(MarketplaceProduct.objects.filter(store=self.store, marketplace="wb").count(), 2)
        product = MarketplaceProduct.objects.get(store=self.store, sku="101")
        self.assertEqual(product.external_ids["source"], "wb_prices_api")
        self.assertEqual(product.last_values["price"], "1000")
        self.assertEqual(
            MarketplaceProductHistory.objects.filter(product=product, operation=operation, file_version=output).count(),
            1,
        )
        conflict_detail = OperationDetailRow.objects.get(operation=operation, product_ref="102")
        self.assertEqual(conflict_detail.reason_code, REASON_SIZE_CONFLICT)
        self.assertFalse(conflict_detail.final_value["upload_ready"])

    def test_permission_and_object_access_required(self):
        User = get_user_model()
        outsider = User.objects.create_user(
            login=f"outsider-{self._testMethodName}",
            password="password",
            display_name="Outsider",
            primary_role=Role.objects.get(code=ROLE_MARKETPLACE_MANAGER),
        )

        with self.assertRaises(PermissionDenied):
            download_wb_prices(
                actor=outsider,
                store=self.store,
                client_factory=FakeClientFactory([[]]),
                secret_resolver=self._secret_resolver,
            )

        permission = Permission.objects.get(code="wb.api.prices.download")
        UserPermissionOverride.objects.create(
            user=self.user,
            store=self.store,
            permission=permission,
            effect=AccessEffect.DENY,
        )
        with self.assertRaises(PermissionDenied):
            download_wb_prices(
                actor=self.user,
                store=self.store,
                client_factory=FakeClientFactory([[]]),
                secret_resolver=self._secret_resolver,
            )

    def test_secret_redaction_in_operation_audit_techlog_and_snapshots(self):
        operation = download_wb_prices(
            actor=self.user,
            store=self.store,
            client_factory=FakeClientFactory([[self._good(201)], []]),
            secret_resolver=self._secret_resolver,
        )
        combined = str(
            {
                "operation": Operation.objects.filter(pk=operation.pk).values(
                    "execution_context",
                    "summary",
                ).get(),
                "audit": list(
                    AuditRecord.objects.filter(operation=operation).values(
                        "safe_message",
                        "before_snapshot",
                        "after_snapshot",
                    )
                ),
                "techlog": list(
                    TechLogRecord.objects.filter(operation=operation).values(
                        "safe_message",
                        "sensitive_details_ref",
                    )
                ),
            },
        )

        self.assertFalse(contains_secret_like(combined))
        self.assertNotIn("task012-local-token-value", combined)
        self.assertNotIn("Authorization", combined)

    def test_active_connection_required(self):
        self.connection.status = ConnectionBlock.Status.CONFIGURED
        self.connection.save(update_fields=["status", "updated_at"])

        with self.assertRaises(PermissionDenied):
            download_wb_prices(
                actor=self.user,
                store=self.store,
                client_factory=FakeClientFactory([[]]),
                secret_resolver=self._secret_resolver,
            )

    def test_api_failure_writes_safe_techlog_and_failed_operation(self):
        class InvalidClient:
            def list_goods_filter(self, *, limit, offset):
                return {"unexpected": []}

        def invalid_factory(*, token, store_scope):
            return InvalidClient()

        with self.assertRaises(Exception):
            download_wb_prices(
                actor=self.user,
                store=self.store,
                client_factory=invalid_factory,
                secret_resolver=self._secret_resolver,
            )

        operation = Operation.objects.get(step_code=OperationStepCode.WB_API_PRICES_DOWNLOAD)
        self.assertEqual(operation.status, "interrupted_failed")
        self.assertEqual(operation.operation_type, OperationType.NOT_APPLICABLE)
        self.assertTrue(TechLogRecord.objects.filter(operation=operation).exists())
        techlog_text = str(TechLogRecord.objects.filter(operation=operation).values())
        self.assertNotIn("task012-local-token-value", techlog_text)
        self.assertFalse(contains_secret_like(techlog_text))

    def test_operation_classifier_contract_rejects_check_process_for_api_step(self):
        run = create_run(
            marketplace=Marketplace.WB,
            module="wb_api",
            mode=OperationMode.API,
            store=self.store,
            initiated_by=self.user,
        )
        operation = Operation(
            marketplace=Marketplace.WB,
            module="wb_api",
            mode=OperationMode.API,
            operation_type=OperationType.CHECK,
            step_code=OperationStepCode.WB_API_PRICES_DOWNLOAD,
            status="created",
            run=run,
            store=self.store,
            initiator_user=self.user,
            logic_version="test",
        )

        with self.assertRaises(ValidationError):
            operation.full_clean()
