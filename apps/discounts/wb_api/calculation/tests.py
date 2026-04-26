from __future__ import annotations

import shutil
import tempfile
from io import BytesIO

from django.contrib.auth import get_user_model
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.test import TestCase, override_settings
from django.utils import timezone
from openpyxl import Workbook, load_workbook

from apps.audit.models import AuditRecord
from apps.discounts.wb_api.promotions.models import (
    WBPromotion,
    WBPromotionExportFile,
    WBPromotionSnapshot,
)
from apps.discounts.wb_api.redaction import contains_secret_like
from apps.discounts.wb_excel.services import calculate as stage1_calculate
from apps.discounts.wb_excel.services import resolve_wb_parameters
from apps.files.models import FileObject
from apps.files.services import create_file_version
from apps.identity_access.models import AccessEffect, Role, StoreAccess
from apps.identity_access.seeds import ROLE_MARKETPLACE_MANAGER, seed_identity_access
from apps.operations.models import (
    Marketplace,
    Operation,
    OperationOutputFile,
    OperationStepCode,
    OperationType,
    OutputKind,
    ProcessStatus,
)
from apps.operations.services import ApiOperationResult, complete_api_operation, create_api_operation, start_operation
from apps.stores.models import StoreAccount
from apps.techlog.models import TechLogRecord

from .services import calculate_wb_api_discounts


class WBApiCalculationTask014Tests(TestCase):
    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls._media_root = tempfile.mkdtemp(prefix="promo-v2-wb-api-calculation-")
        cls.override = override_settings(MEDIA_ROOT=cls._media_root)
        cls.override.enable()

    @classmethod
    def tearDownClass(cls):
        cls.override.disable()
        shutil.rmtree(cls._media_root, ignore_errors=True)
        super().tearDownClass()

    def setUp(self):
        seed_identity_access()
        role = Role.objects.get(code=ROLE_MARKETPLACE_MANAGER)
        User = get_user_model()
        self.user = User.objects.create_user(
            login=f"manager-{self._testMethodName}",
            password="password",
            display_name="Manager",
            primary_role=role,
        )
        self.store = StoreAccount.objects.create(
            name=f"WB API Calc Store {self._testMethodName}",
            marketplace=StoreAccount.Marketplace.WB,
        )
        StoreAccess.objects.create(
            user=self.user,
            store=self.store,
            access_level=StoreAccess.AccessLevel.WORK,
            effect=AccessEffect.ALLOW,
        )

    def _xlsx_version(self, *, rows: list[list], scenario: str, name: str, logical_name: str):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Цены" if "price" in logical_name else "Акция"
        for row in rows:
            sheet.append(row)
        raw = workbook.create_sheet("_api_raw")
        raw.append(["safe"])
        raw.append(["no-token"])
        buffer = BytesIO()
        workbook.save(buffer)
        workbook.close()
        return create_file_version(
            store=self.store,
            uploaded_by=self.user,
            uploaded_file=ContentFile(buffer.getvalue(), name=name),
            scenario=scenario,
            kind=FileObject.Kind.OUTPUT,
            logical_name=logical_name,
            module="wb_api",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def _api_operation(self, *, step_code: str):
        operation = create_api_operation(
            marketplace=Marketplace.WB,
            store=self.store,
            initiator_user=self.user,
            step_code=step_code,
            logic_version=f"{step_code}-test",
            execution_context={"step_code": step_code},
        )
        return start_operation(operation)

    def _complete_operation(self, operation, *, summary: dict, output_file=None, status=ProcessStatus.COMPLETED_SUCCESS):
        return complete_api_operation(
            operation,
            result=ApiOperationResult(
                summary=summary,
                status=status,
                output_file_version=output_file,
                output_kind=OutputKind.OUTPUT_WORKBOOK,
            ),
        )

    def _price_basis(self, rows: list[list], *, checksum="price-checksum"):
        operation = self._api_operation(step_code=OperationStepCode.WB_API_PRICES_DOWNLOAD)
        file_version = self._xlsx_version(
            rows=[["Артикул WB", "Текущая цена", "Новая скидка", "Не менять"], *rows],
            scenario=FileObject.Scenario.WB_DISCOUNTS_API_PRICE_EXPORT,
            name="wb-api-prices.xlsx",
            logical_name="api_price",
        )
        operation = self._complete_operation(
            operation,
            output_file=file_version,
            summary={"safe_snapshot": {"source_checksum": checksum}},
        )
        return operation, file_version

    def _promo_basis(self, rows: list[list], *, promotion_id=101, checksum="promo-checksum"):
        operation = self._api_operation(step_code=OperationStepCode.WB_API_PROMOTIONS_DOWNLOAD)
        file_version = self._xlsx_version(
            rows=[
                [
                    "Артикул WB",
                    "Плановая цена для акции",
                    "Загружаемая скидка для участия в акции",
                ],
                *rows,
            ],
            scenario=FileObject.Scenario.WB_DISCOUNTS_API_PROMOTION_EXPORT,
            name=f"wb_promo_{promotion_id}.xlsx",
            logical_name=f"wb_promo_{promotion_id}",
        )
        now = timezone.now()
        snapshot = WBPromotionSnapshot.objects.create(
            operation=operation,
            store=self.store,
            fetched_at=now,
            api_window_start=now,
            api_window_end=now,
            current_filter_timestamp=now,
            raw_response_safe_snapshot={"source_checksum": checksum},
            promotions_count=1,
            current_promotions_count=1,
            regular_current_promotions_count=1,
            promotion_products_count=len(rows),
        )
        promotion = WBPromotion.objects.create(
            store=self.store,
            wb_promotion_id=promotion_id,
            name="Current regular",
            type="regular",
            start_datetime=now,
            end_datetime=now,
            is_current_at_fetch=True,
            last_seen_at=now,
            snapshot_ref=snapshot,
        )
        OperationOutputFile.objects.create(
            operation=operation,
            file_version=file_version,
            output_kind=OutputKind.PROMOTION_EXPORT,
        )
        WBPromotionExportFile.objects.create(
            promotion=promotion,
            operation=operation,
            file_version=file_version,
        )
        operation = self._complete_operation(
            operation,
            summary={
                "safe_snapshot": {"source_checksum": checksum},
                "wb_promotion_snapshot_id": snapshot.pk,
            },
        )
        return operation, file_version

    def test_api_calculation_matches_stage1_logic_and_writes_result_excel_only_new_discount(self):
        price_operation, price_file = self._price_basis(
            [["123", "1000", 1, "left-a"], ["456", "1000", 2, "=1+1"], ["789", "100", 3, "left-c"]],
        )
        promo_operation, promo_file = self._promo_basis(
            [["123", "333.10", "80"], ["123", "500", "60"], ["456", "100", "90"]],
        )
        stage1_result = stage1_calculate(price_file, [promo_file], resolve_wb_parameters(self.store))

        operation = calculate_wb_api_discounts(
            actor=self.user,
            store=self.store,
            price_operation=price_operation,
            promotion_operation=promo_operation,
        )

        self.assertEqual(operation.marketplace, Marketplace.WB)
        self.assertEqual(operation.step_code, OperationStepCode.WB_API_DISCOUNT_CALCULATION)
        self.assertEqual(operation.operation_type, OperationType.NOT_APPLICABLE)
        self.assertEqual(operation.status, ProcessStatus.COMPLETED_SUCCESS)
        self.assertEqual(operation.parameter_snapshots.count(), 3)
        self.assertEqual(operation.summary["basis"]["price_basis"]["operation_id"], price_operation.pk)
        self.assertEqual(operation.summary["basis"]["promotion_basis"]["operation_id"], promo_operation.pk)
        self.assertEqual(operation.summary["calculated_rows"], len(stage1_result.final_discounts_by_row))

        api_details = {
            detail.row_no: detail.final_value["final_discount"]
            for detail in operation.detail_rows.filter(row_status="ok")
        }
        self.assertEqual(api_details, stage1_result.final_discounts_by_row)
        self.assertEqual(
            set(operation.detail_rows.filter(row_status="ok").values_list("reason_code", flat=True)),
            {"wb_api_calculated_from_api_sources"},
        )

        output = operation.output_files.get().file_version
        self.assertEqual(output.file.scenario, FileObject.Scenario.WB_DISCOUNTS_API_RESULT_EXCEL)
        with default_storage.open(output.storage_path, "rb") as handle:
            workbook = load_workbook(handle, data_only=False)
        sheet = workbook[workbook.sheetnames[0]]
        self.assertEqual(sheet["A2"].value, "123")
        self.assertEqual(sheet["B2"].value, "1000")
        self.assertEqual(sheet["C2"].value, 50)
        self.assertEqual(sheet["D2"].value, "left-a")
        self.assertEqual(sheet["C3"].value, 55)
        self.assertEqual(sheet["D3"].value, "=1+1")
        self.assertEqual(sheet["C4"].value, 55)
        self.assertEqual(sheet["D4"].value, "left-c")
        self.assertIn("_api_raw", workbook.sheetnames)
        workbook.close()

    def test_latest_basis_is_selected_and_recalculation_creates_new_operation_and_file_version(self):
        self._price_basis([["old", "1000", ""]], checksum="old-price")
        latest_price_operation, _latest_price = self._price_basis([["new", "1000", ""]], checksum="new-price")
        promo_operation, _promo_file = self._promo_basis([["new", "900", "15"]])

        first = calculate_wb_api_discounts(actor=self.user, store=self.store)
        second = calculate_wb_api_discounts(actor=self.user, store=self.store)

        self.assertEqual(first.summary["basis"]["price_basis"]["operation_id"], latest_price_operation.pk)
        self.assertEqual(first.summary["basis"]["promotion_basis"]["operation_id"], promo_operation.pk)
        self.assertNotEqual(first.pk, second.pk)
        self.assertNotEqual(first.output_files.get().file_version_id, second.output_files.get().file_version_id)

    def test_errors_block_result_output_and_upload_basis(self):
        price_operation, _price_file = self._price_basis([["123", "100", ""]])
        promo_operation, _promo_file = self._promo_basis([["123", "150", "10"]])

        operation = calculate_wb_api_discounts(
            actor=self.user,
            store=self.store,
            price_operation=price_operation,
            promotion_operation=promo_operation,
        )

        self.assertEqual(operation.status, ProcessStatus.COMPLETED_WITH_ERROR)
        self.assertEqual(operation.error_count, 1)
        self.assertTrue(operation.summary["upload_blocked"])
        self.assertFalse(operation.output_files.exists())
        self.assertTrue(operation.detail_rows.filter(reason_code="wb_discount_out_of_range").exists())

    def test_secret_like_values_are_absent_from_operation_audit_techlog_and_result_file(self):
        price_operation, _price_file = self._price_basis([["123", "1000", ""]])
        promo_operation, _promo_file = self._promo_basis([["123", "900", "15"]])

        operation = calculate_wb_api_discounts(
            actor=self.user,
            store=self.store,
            price_operation=price_operation,
            promotion_operation=promo_operation,
        )
        output = operation.output_files.get().file_version
        with default_storage.open(output.storage_path, "rb") as handle:
            workbook_bytes = handle.read()
        combined = str(
            {
                "operation": Operation.objects.filter(pk=operation.pk).values("execution_context", "summary").get(),
                "audit": list(
                    AuditRecord.objects.filter(operation=operation).values(
                        "safe_message",
                        "before_snapshot",
                        "after_snapshot",
                    )
                ),
                "techlog": list(
                    TechLogRecord.objects.filter(operation=operation).values(
                        "safe_message",
                        "sensitive_details_ref",
                    )
                ),
                "file_text": workbook_bytes.decode("latin1", errors="ignore"),
            },
        )

        self.assertFalse(contains_secret_like(combined))
        self.assertNotIn("Bearer", combined)
