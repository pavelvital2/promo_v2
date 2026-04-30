from __future__ import annotations

from django.core.exceptions import PermissionDenied, ValidationError
from django.core.files.storage import default_storage
from django.test import SimpleTestCase, TestCase
from openpyxl import load_workbook

from apps.discounts.ozon_api.actions import (
    ELASTIC_ACTION_TYPE,
    SELECTED_ACTION_METADATA_KEY,
    download_ozon_actions,
    filter_elastic_actions,
    get_selected_elastic_action_basis,
    select_elastic_action,
)
from apps.discounts.ozon_api.client import (
    OZON_ACTION_CANDIDATES_PATH,
    OZON_ACTION_PRODUCTS_PATH,
    OZON_ACTION_PRODUCTS_ACTIVATE_PATH,
    OZON_ACTION_PRODUCTS_DEACTIVATE_PATH,
    OZON_ACTIONS_PATH,
    OZON_PRODUCT_INFO_LIST_PATH,
    OZON_PRODUCT_INFO_STOCKS_PATH,
    OzonApiAuthError,
    OzonApiClient,
    OzonApiCredentials,
    OzonApiInvalidResponseError,
    OzonApiPolicy,
    OzonApiRateLimitError,
    OzonApiTemporaryError,
)
from apps.discounts.ozon_api.calculation import calculate_elastic_result
from apps.discounts.ozon_api.review import (
    accept_elastic_result,
    decline_elastic_result,
    is_upload_allowed_by_review,
)
from apps.discounts.ozon_api.upload import (
    OzonElasticDeactivateUnconfirmed,
    OzonElasticDriftDetected,
    confirm_deactivate_group,
    deactivate_confirmation_preview,
    upload_elastic_result,
)
from apps.discounts.ozon_api.products import (
    ACTIVE_SOURCE_GROUP,
    CANDIDATE_SOURCE_GROUP,
    build_latest_product_source_basis,
    download_active_products,
    download_candidate_products,
)
from apps.discounts.ozon_api.product_data import download_product_data
from apps.discounts.ozon_excel.services import decide_row as decide_excel_row
from apps.discounts.ozon_shared.calculation import decide_ozon_row, parse_decimal
from apps.files.models import FileObject
from apps.audit.models import AuditActionCode, AuditRecord
from apps.identity_access.models import Role, StoreAccess, User
from apps.identity_access.seeds import ROLE_LOCAL_ADMIN, ROLE_MARKETPLACE_MANAGER, ROLE_OWNER, seed_identity_access
from apps.operations.models import Marketplace, OperationMode, OperationModule, OperationStepCode, OperationType
from apps.operations.models import Operation
from apps.stores.models import ConnectionBlock, StoreAccount
from apps.stores.services import OZON_API_CONNECTION_TYPE, OZON_API_MODULE


CLIENT_ID = "123456"
API_KEY = "ozon-api-key-abcdefghijklmnopqrstuvwxyz"


class FakeResponse:
    def __init__(self, status_code: int, payload=None, *, json_error: bool = False):
        self.status_code = status_code
        self.payload = payload
        self.json_error = json_error

    def json(self):
        if self.json_error:
            raise ValueError("invalid json")
        return self.payload


class RecordingSession:
    def __init__(self, responses):
        self.responses = list(responses)
        self.calls = []

    def get(self, url, *, params=None, headers=None, timeout=None):
        self.calls.append(
            {
                "method": "GET",
                "url": url,
                "params": params or {},
                "headers": headers or {},
                "timeout": timeout,
            },
        )
        response = self.responses.pop(0)
        if isinstance(response, BaseException):
            raise response
        return response

    def post(self, url, *, json=None, headers=None, timeout=None):
        self.calls.append(
            {
                "method": "POST",
                "url": url,
                "json": json or {},
                "headers": headers or {},
                "timeout": timeout,
            },
        )
        response = self.responses.pop(0)
        if isinstance(response, BaseException):
            raise response
        return response


class OzonApiClientTests(SimpleTestCase):
    def _client(self, session: RecordingSession) -> OzonApiClient:
        return OzonApiClient(
            credentials=OzonApiCredentials(client_id=CLIENT_ID, api_key=API_KEY),
            base_url="https://ozon.example",
            session=session,
            policy=OzonApiPolicy(
                max_read_retries=2,
                backoff_seconds=0,
                min_interval_seconds=0,
            ),
            store_scope="STORE-000001",
        )

    def test_connection_check_uses_only_get_actions_and_validates_result(self):
        session = RecordingSession([FakeResponse(200, {"result": []})])

        result = self._client(session).check_connection()

        self.assertEqual(result, {"result": []})
        self.assertEqual(len(session.calls), 1)
        call = session.calls[0]
        self.assertEqual(call["method"], "GET")
        self.assertTrue(call["url"].endswith(OZON_ACTIONS_PATH))
        self.assertEqual(call["params"], {})
        self.assertEqual(call["headers"]["Client-Id"], CLIENT_ID)
        self.assertEqual(call["headers"]["Api-Key"], API_KEY)

    def test_connection_check_maps_auth_rate_temporary_and_invalid_response(self):
        cases = (
            ([FakeResponse(401, {})], OzonApiAuthError),
            ([FakeResponse(403, {})], OzonApiAuthError),
            ([FakeResponse(429, {}), FakeResponse(429, {}), FakeResponse(429, {})], OzonApiRateLimitError),
            ([FakeResponse(500, {}), FakeResponse(502, {}), FakeResponse(503, {})], OzonApiTemporaryError),
            ([TimeoutError("timeout"), TimeoutError("timeout"), TimeoutError("timeout")], OzonApiTemporaryError),
            ([OSError("network"), OSError("network"), OSError("network")], OzonApiTemporaryError),
            ([FakeResponse(200, json_error=True)], OzonApiInvalidResponseError),
            ([FakeResponse(200, {"items": []})], OzonApiInvalidResponseError),
            ([FakeResponse(200, [])], OzonApiInvalidResponseError),
        )

        for responses, error_class in cases:
            with self.subTest(error_class=error_class, responses=responses):
                session = RecordingSession(responses)
                with self.assertRaises(error_class):
                    self._client(session).check_connection()
                self.assertTrue(session.calls)
                self.assertTrue(all(call["method"] == "GET" for call in session.calls))
                self.assertTrue(all(call["url"].endswith(OZON_ACTIONS_PATH) for call in session.calls))

    def test_policy_defaults_match_stage_2_2_baseline(self):
        policy = OzonApiPolicy()

        self.assertEqual(policy.read_page_size, 100)
        self.assertEqual(policy.write_batch_size, 100)
        self.assertEqual(policy.min_interval_seconds, 0.5)

    def test_list_actions_uses_get_actions_with_page_defaults(self):
        session = RecordingSession([FakeResponse(200, {"result": []})])

        result = self._client(session).list_actions(limit=100, offset=0)

        self.assertEqual(result, {"result": []})
        self.assertEqual(session.calls[0]["method"], "GET")
        self.assertTrue(session.calls[0]["url"].endswith(OZON_ACTIONS_PATH))
        self.assertEqual(session.calls[0]["params"], {"limit": 100, "offset": 0})

    def test_action_product_reads_use_post_with_safe_body_and_retry_transient_only(self):
        session = RecordingSession(
            [
                FakeResponse(429, {}),
                FakeResponse(
                    200,
                    {
                        "result": {
                            "products": [{"id": 101, "min_action_price": 100, "max_action_price": 200}],
                            "total": 1,
                            "last_id": "",
                        }
                    },
                ),
            ]
        )

        result = self._client(session).list_action_products(action_id="elastic-1", limit=100, offset=0)

        self.assertEqual(result["result"]["total"], 1)
        self.assertEqual(len(session.calls), 2)
        self.assertTrue(all(call["method"] == "POST" for call in session.calls))
        self.assertTrue(all(call["url"].endswith(OZON_ACTION_PRODUCTS_PATH) for call in session.calls))
        self.assertEqual(session.calls[-1]["json"], {"action_id": "elastic-1", "limit": 100, "offset": 0})

        invalid = RecordingSession([FakeResponse(400, {})])
        with self.assertRaises(OzonApiInvalidResponseError):
            self._client(invalid).list_action_candidates(action_id="elastic-1", limit=100)
        self.assertEqual(len(invalid.calls), 1)
        self.assertTrue(invalid.calls[0]["url"].endswith(OZON_ACTION_CANDIDATES_PATH))

    def test_product_info_and_stocks_use_default_chunks_and_read_retry(self):
        session = RecordingSession(
            [
                FakeResponse(429, {}),
                FakeResponse(200, {"result": {"items": [{"id": "101", "min_price": "123"}]}}),
                FakeResponse(500, {}),
                FakeResponse(200, {"result": {"items": [{"product_id": "101", "stocks": [{"present": 3}]}]}}),
            ]
        )
        client = self._client(session)

        info = client.product_info_list(product_ids=["101"])
        stocks = client.product_info_stocks(product_ids=["101"])

        self.assertEqual(info["result"]["items"][0]["min_price"], "123")
        self.assertEqual(stocks["result"]["items"][0]["stocks"][0]["present"], 3)
        self.assertEqual(len(session.calls), 4)
        self.assertTrue(all(call["method"] == "POST" for call in session.calls))
        self.assertTrue(session.calls[0]["url"].endswith(OZON_PRODUCT_INFO_LIST_PATH))
        self.assertTrue(session.calls[2]["url"].endswith(OZON_PRODUCT_INFO_STOCKS_PATH))
        self.assertEqual(session.calls[1]["json"], {"product_id": ["101"]})

    def test_write_activate_deactivate_use_actions_endpoints_without_retry(self):
        session = RecordingSession(
            [
                FakeResponse(429, {}),
                FakeResponse(200, {"result": {}}),
            ]
        )
        client = self._client(session)

        with self.assertRaises(OzonApiRateLimitError):
            client.activate_action_products(
                action_id="elastic-1",
                products=[{"product_id": "101", "action_price": "100"}],
            )
        result = client.deactivate_action_products(
            action_id="elastic-1",
            products=[{"product_id": "101"}],
        )

        self.assertEqual(result, {"result": {}})
        self.assertEqual(len(session.calls), 2)
        self.assertTrue(session.calls[0]["url"].endswith(OZON_ACTION_PRODUCTS_ACTIVATE_PATH))
        self.assertTrue(session.calls[1]["url"].endswith(OZON_ACTION_PRODUCTS_DEACTIVATE_PATH))
        self.assertEqual(
            session.calls[0]["json"],
            {"action_id": "elastic-1", "products": [{"product_id": "101", "action_price": "100"}]},
        )


class FakeActionsClient:
    responses = []
    credentials_seen = None
    store_scope_seen = None

    def __init__(self, *, credentials, store_scope, **kwargs):
        type(self).credentials_seen = credentials
        type(self).store_scope_seen = store_scope
        self.policy = OzonApiPolicy(max_read_retries=2, backoff_seconds=0, min_interval_seconds=0)

    def list_actions(self, *, limit=None, offset=0):
        return type(self).responses.pop(0)


class FakeProductsClient:
    active_responses = []
    candidate_responses = []
    calls = []

    def __init__(self, *, credentials, store_scope, **kwargs):
        self.policy = OzonApiPolicy(max_read_retries=2, backoff_seconds=0, min_interval_seconds=0)

    def list_action_products(self, *, action_id, limit=None, offset=0, last_id=""):
        type(self).calls.append(
            {
                "endpoint": "/v1/actions/products",
                "action_id": action_id,
                "limit": limit,
                "offset": offset,
                "last_id": last_id,
            }
        )
        return type(self).active_responses.pop(0)

    def list_action_candidates(self, *, action_id, limit=None, offset=0, last_id=""):
        type(self).calls.append(
            {
                "endpoint": "/v1/actions/candidates",
                "action_id": action_id,
                "limit": limit,
                "offset": offset,
                "last_id": last_id,
            }
        )
        return type(self).candidate_responses.pop(0)


class FakeProductDataClient:
    product_info_responses = []
    stock_responses = []
    calls = []

    def __init__(self, *, credentials, store_scope, **kwargs):
        self.policy = OzonApiPolicy(max_read_retries=2, backoff_seconds=0, min_interval_seconds=0)

    def product_info_list(self, *, product_ids):
        type(self).calls.append(
            {
                "endpoint": "/v3/product/info/list",
                "product_ids": list(product_ids),
                "limit": self.policy.read_page_size,
            }
        )
        return type(self).product_info_responses.pop(0)

    def product_info_stocks(self, *, product_ids):
        type(self).calls.append(
            {
                "endpoint": "/v4/product/info/stocks",
                "product_ids": list(product_ids),
                "limit": self.policy.read_page_size,
            }
        )
        return type(self).stock_responses.pop(0)


class FakeUploadClient:
    calls = []
    actions_response = None
    actions_responses = []
    active_response = None
    active_membership_responses = []
    candidate_response = None
    candidate_membership_responses = []
    product_info_response = None
    product_info_responses = []
    stock_response = None
    stock_responses = []
    activate_responses = []
    deactivate_responses = []
    policy = OzonApiPolicy(max_read_retries=2, backoff_seconds=0, min_interval_seconds=0)

    def __init__(self, *, credentials, store_scope, **kwargs):
        self.policy = type(self).policy

    def list_actions(self, *, limit=None, offset=0):
        type(self).calls.append({"endpoint": "/v1/actions", "limit": limit, "offset": offset})
        if type(self).actions_responses:
            return type(self).actions_responses.pop(0)
        return type(self).actions_response

    def list_action_products(self, *, action_id, limit=None, offset=0, last_id=""):
        type(self).calls.append(
            {"endpoint": "/v1/actions/products", "action_id": action_id, "limit": limit, "offset": offset}
        )
        if type(self).active_membership_responses:
            return type(self).active_membership_responses.pop(0)
        return type(self).active_response

    def list_action_candidates(self, *, action_id, limit=None, offset=0, last_id=""):
        type(self).calls.append(
            {"endpoint": "/v1/actions/candidates", "action_id": action_id, "limit": limit, "offset": offset}
        )
        if type(self).candidate_membership_responses:
            return type(self).candidate_membership_responses.pop(0)
        return type(self).candidate_response

    def product_info_list(self, *, product_ids):
        type(self).calls.append({"endpoint": "/v3/product/info/list", "product_ids": list(product_ids)})
        if type(self).product_info_responses:
            return type(self).product_info_responses.pop(0)
        return type(self).product_info_response

    def product_info_stocks(self, *, product_ids):
        type(self).calls.append({"endpoint": "/v4/product/info/stocks", "product_ids": list(product_ids)})
        if type(self).stock_responses:
            return type(self).stock_responses.pop(0)
        return type(self).stock_response

    def activate_action_products(self, *, action_id, products):
        type(self).calls.append(
            {"endpoint": "/v1/actions/products/activate", "action_id": action_id, "products": list(products)}
        )
        response = type(self).activate_responses.pop(0)
        if isinstance(response, BaseException):
            raise response
        return response

    def deactivate_action_products(self, *, action_id, products):
        type(self).calls.append(
            {"endpoint": "/v1/actions/products/deactivate", "action_id": action_id, "products": list(products)}
        )
        response = type(self).deactivate_responses.pop(0)
        if isinstance(response, BaseException):
            raise response
        return response


def fake_secret_resolver(_protected_secret_ref):
    return OzonApiCredentials(client_id=CLIENT_ID, api_key=API_KEY)


class OzonActionsDownloadTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        seed_identity_access()
        cls.manager_role = Role.objects.get(code=ROLE_MARKETPLACE_MANAGER)

    def setUp(self):
        self.store = StoreAccount.objects.create(
            name=f"Ozon store {self._testMethodName}",
            marketplace=StoreAccount.Marketplace.OZON,
        )
        self.actor = User.objects.create_user(
            login=f"manager-{self._testMethodName}",
            password="pass",
            display_name="Manager",
            primary_role=self.manager_role,
        )
        self.owner = User.objects.create_user(
            login=f"owner-{self._testMethodName}",
            password="pass",
            display_name="Owner",
            primary_role=Role.objects.get(code=ROLE_OWNER),
        )
        StoreAccess.objects.create(user=self.actor, store=self.store, access_level=StoreAccess.AccessLevel.WORK)
        self.connection = ConnectionBlock.objects.create(
            store=self.store,
            module=OZON_API_MODULE,
            connection_type=OZON_API_CONNECTION_TYPE,
            status=ConnectionBlock.Status.ACTIVE,
            protected_secret_ref="env://OZON_SECRET",
            metadata={},
        )

    def _elastic(self, action_id="elastic-1", title="Эластичный бустинг апрель"):
        return {
            "id": action_id,
            "title": title,
            "action_type": ELASTIC_ACTION_TYPE,
            "status": "ACTIVE",
            "Client-Id": CLIENT_ID,
            "Api-Key": API_KEY,
        }

    def test_filtering_elastic_non_elastic_and_ambiguous_actions(self):
        elastic, ambiguous, non_elastic = filter_elastic_actions(
            [
                self._elastic(),
                {"id": "ambiguous-1", "title": "Эластичный бустинг без типа", "action_type": "OTHER"},
                {"id": "ambiguous-2", "title": "Other title", "action_type": ELASTIC_ACTION_TYPE},
                {"id": "regular-1", "title": "Regular", "action_type": "OTHER"},
            ]
        )

        self.assertEqual([item["action_id"] for item in elastic], ["elastic-1"])
        self.assertEqual({item["action_id"] for item in ambiguous}, {"ambiguous-1", "ambiguous-2"})
        self.assertEqual([item["action_id"] for item in non_elastic], ["regular-1"])

    def test_actions_download_operation_classifier_summary_and_safe_snapshot(self):
        FakeActionsClient.responses = [
            {
                "result": [
                    self._elastic("elastic-101"),
                    {"id": "regular-1", "title": "Regular", "action_type": "OTHER"},
                    {"id": "ambiguous-1", "title": "Эластичный бустинг", "action_type": "OTHER"},
                ]
            }
        ]

        operation = download_ozon_actions(
            actor=self.actor,
            store=self.store,
            client_factory=FakeActionsClient,
            secret_resolver=fake_secret_resolver,
        )

        self.assertEqual(operation.marketplace, Marketplace.OZON)
        self.assertEqual(operation.module, OperationModule.OZON_API)
        self.assertEqual(operation.mode, OperationMode.API)
        self.assertEqual(operation.operation_type, OperationType.NOT_APPLICABLE)
        self.assertEqual(operation.step_code, OperationStepCode.OZON_API_ACTIONS_DOWNLOAD)
        self.assertEqual(operation.summary["actions_count"], 3)
        self.assertEqual(operation.summary["elastic_actions_count"], 1)
        self.assertEqual(operation.summary["ambiguous_actions_count"], 1)
        self.assertEqual(operation.summary["non_elastic_actions_count"], 1)
        self.assertNotIn("result_code", operation.summary)
        self.assertEqual(operation.summary["selection_basis"]["available_action_ids"], ["elastic-101"])
        dumped = str(operation.summary)
        self.assertNotIn(CLIENT_ID, dumped)
        self.assertNotIn(API_KEY, dumped)
        self.assertNotIn("Client-Id", dumped)
        self.assertNotIn("Api-Key", dumped)
        self.assertEqual(operation.detail_rows.count(), 3)
        self.assertEqual(set(operation.detail_rows.values_list("reason_code", flat=True)), {""})

    def test_download_requires_active_connection_permission_and_object_access(self):
        self.connection.status = ConnectionBlock.Status.CONFIGURED
        self.connection.save()
        with self.assertRaises(PermissionDenied):
            download_ozon_actions(
                actor=self.actor,
                store=self.store,
                client_factory=FakeActionsClient,
                secret_resolver=fake_secret_resolver,
            )

        self.connection.status = ConnectionBlock.Status.ACTIVE
        self.connection.save()
        denied = User.objects.create_user(
            login=f"denied-{self._testMethodName}",
            password="pass",
            display_name="Denied",
            primary_role=self.manager_role,
        )
        with self.assertRaises(PermissionDenied):
            download_ozon_actions(
                actor=denied,
                store=self.store,
                client_factory=FakeActionsClient,
                secret_resolver=fake_secret_resolver,
            )

    def test_selected_action_id_persisted_as_store_connection_basis(self):
        FakeActionsClient.responses = [{"result": [self._elastic("elastic-store-a"), self._elastic("elastic-store-b")]}]
        operation = download_ozon_actions(
            actor=self.actor,
            store=self.store,
            client_factory=FakeActionsClient,
            secret_resolver=fake_secret_resolver,
        )

        basis = select_elastic_action(actor=self.actor, store=self.store, action_id="elastic-store-b")

        self.assertEqual(basis["action_id"], "elastic-store-b")
        self.assertEqual(basis["source_operation_id"], operation.pk)
        self.connection.refresh_from_db()
        self.assertEqual(
            self.connection.metadata[SELECTED_ACTION_METADATA_KEY]["action_id"],
            "elastic-store-b",
        )
        self.assertEqual(get_selected_elastic_action_basis(self.store)["action_id"], "elastic-store-b")

        with self.assertRaises(ValidationError):
            select_elastic_action(actor=self.actor, store=self.store, action_id="hard-coded-global-action")

    def _select_action(self, action_id="elastic-products"):
        FakeActionsClient.responses = [{"result": [self._elastic(action_id)]}]
        download_ozon_actions(
            actor=self.actor,
            store=self.store,
            client_factory=FakeActionsClient,
            secret_resolver=fake_secret_resolver,
        )
        select_elastic_action(actor=self.actor, store=self.store, action_id=action_id)

    def test_active_products_download_uses_saved_action_id_and_paginates(self):
        self._select_action("elastic-active")
        FakeProductsClient.calls = []
        FakeProductsClient.active_responses = [
            {
                "result": {
                    "products": [
                        {"id": "101", "min_action_price": 100, "max_action_price": 200},
                    ],
                    "total": 2,
                    "last_id": "page-2",
                }
            },
            {
                "result": {
                    "products": [
                        {"id": "102", "min_action_price": 110, "max_action_price": 210},
                    ],
                    "total": 2,
                    "last_id": "",
                }
            },
        ]

        operation = download_active_products(
            actor=self.actor,
            store=self.store,
            client_factory=FakeProductsClient,
            secret_resolver=fake_secret_resolver,
        )

        self.assertEqual(operation.step_code, OperationStepCode.OZON_API_ELASTIC_ACTIVE_PRODUCTS_DOWNLOAD)
        self.assertEqual(operation.summary["action_id"], "elastic-active")
        self.assertEqual(operation.summary["source_group"], ACTIVE_SOURCE_GROUP)
        self.assertEqual(operation.summary["products_count"], 2)
        self.assertNotIn("result_code", operation.summary)
        self.assertEqual(
            list(operation.detail_rows.order_by("row_no").values_list("reason_code", flat=True)),
            ["", ""],
        )
        self.assertEqual(FakeProductsClient.calls[0]["limit"], 100)
        self.assertEqual(FakeProductsClient.calls[0]["action_id"], "elastic-active")
        self.assertEqual(FakeProductsClient.calls[1]["last_id"], "page-2")
        self.assertTrue(all(call["endpoint"] == "/v1/actions/products" for call in FakeProductsClient.calls))

    def test_product_download_failure_uses_closed_api_error_code(self):
        self._select_action("elastic-invalid-response")
        FakeProductsClient.calls = []
        FakeProductsClient.active_responses = [{"items": []}]

        with self.assertRaises(OzonApiInvalidResponseError):
            download_active_products(
                actor=self.actor,
                store=self.store,
                client_factory=FakeProductsClient,
                secret_resolver=fake_secret_resolver,
            )

        operation = self.store.operations.order_by("-id").first()
        self.assertEqual(operation.summary["result_code"], "ozon_api_response_invalid")

    def test_candidate_products_download_paginates_and_uses_read_only_endpoint(self):
        self._select_action("elastic-candidates")
        FakeProductsClient.calls = []
        FakeProductsClient.candidate_responses = [
            {
                "result": {
                    "products": [
                        {"id": "201", "min_action_price": 100, "max_action_price": 200},
                    ],
                    "total": 2,
                    "last_id": "page-2",
                }
            },
            {
                "result": {
                    "products": [
                        {"id": "202", "min_action_price": 110, "max_action_price": 210},
                    ],
                    "total": 2,
                    "last_id": "",
                }
            },
        ]

        operation = download_candidate_products(
            actor=self.actor,
            store=self.store,
            client_factory=FakeProductsClient,
            secret_resolver=fake_secret_resolver,
        )

        self.assertEqual(operation.step_code, OperationStepCode.OZON_API_ELASTIC_CANDIDATE_PRODUCTS_DOWNLOAD)
        self.assertEqual(operation.summary["source_group"], CANDIDATE_SOURCE_GROUP)
        self.assertEqual(operation.summary["products_count"], 2)
        self.assertTrue(all(call["endpoint"] == "/v1/actions/candidates" for call in FakeProductsClient.calls))
        self.assertFalse(any("activate" in call["endpoint"] or "deactivate" in call["endpoint"] for call in FakeProductsClient.calls))

    def test_empty_groups_are_persisted_as_successful_snapshots(self):
        self._select_action("elastic-empty")
        FakeProductsClient.calls = []
        FakeProductsClient.active_responses = [{"result": {"products": [], "total": 0, "last_id": ""}}]
        FakeProductsClient.candidate_responses = [{"result": {"products": [], "total": 0, "last_id": ""}}]

        active = download_active_products(
            actor=self.actor,
            store=self.store,
            client_factory=FakeProductsClient,
            secret_resolver=fake_secret_resolver,
        )
        candidate = download_candidate_products(
            actor=self.actor,
            store=self.store,
            client_factory=FakeProductsClient,
            secret_resolver=fake_secret_resolver,
        )

        self.assertEqual(active.summary["products_count"], 0)
        self.assertEqual(candidate.summary["products_count"], 0)
        basis = build_latest_product_source_basis(store=self.store, action_id="elastic-empty")
        self.assertEqual(basis["rows_count"], 0)
        self.assertEqual(basis["missing_groups"], [])

    def test_missing_selected_or_non_elastic_action_blocks_product_download(self):
        FakeProductsClient.active_responses = [{"result": {"products": [], "total": 0}}]
        with self.assertRaises(ValidationError):
            download_active_products(
                actor=self.actor,
                store=self.store,
                client_factory=FakeProductsClient,
                secret_resolver=fake_secret_resolver,
            )

        self.connection.metadata[SELECTED_ACTION_METADATA_KEY] = {
            "action_id": "regular-action",
            "action": {"action_id": "regular-action", "title": "Regular", "action_type": "OTHER"},
        }
        self.connection.save(update_fields=["metadata", "updated_at"])
        with self.assertRaises(ValidationError):
            download_candidate_products(
                actor=self.actor,
                store=self.store,
                client_factory=FakeProductsClient,
                secret_resolver=fake_secret_resolver,
            )

    def test_missing_elastic_fields_use_safe_diagnostic_code_without_raw_response(self):
        self._select_action("elastic-missing-fields")
        FakeProductsClient.calls = []
        FakeProductsClient.active_responses = [
            {
                "result": {
                    "products": [
                        {
                            "id": "301",
                            "Client-Id": CLIENT_ID,
                            "Api-Key": API_KEY,
                        }
                    ],
                    "total": 1,
                }
            }
        ]

        operation = download_active_products(
            actor=self.actor,
            store=self.store,
            client_factory=FakeProductsClient,
            secret_resolver=fake_secret_resolver,
        )

        self.assertEqual(operation.summary["missing_elastic_fields_count"], 1)
        row = operation.detail_rows.get()
        self.assertEqual(row.reason_code, "ozon_api_missing_elastic_fields")
        dumped = str(operation.summary) + str(row.final_value)
        self.assertNotIn(CLIENT_ID, dumped)
        self.assertNotIn(API_KEY, dumped)
        self.assertNotIn("Client-Id", dumped)
        self.assertNotIn("Api-Key", dumped)

    def test_active_candidate_collision_basis_is_merged_with_visible_details(self):
        self._select_action("elastic-collision")
        FakeProductsClient.calls = []
        FakeProductsClient.active_responses = [
            {
                "result": {
                    "products": [
                        {"id": "401", "min_action_price": 100, "max_action_price": 200},
                        {"id": "402", "min_action_price": 120, "max_action_price": 220},
                    ],
                    "total": 2,
                }
            }
        ]
        FakeProductsClient.candidate_responses = [
            {
                "result": {
                    "products": [
                        {"id": "401", "min_action_price": 100, "max_action_price": 200},
                        {"id": "403", "min_action_price": 130, "max_action_price": 230},
                    ],
                    "total": 2,
                }
            }
        ]

        download_active_products(
            actor=self.actor,
            store=self.store,
            client_factory=FakeProductsClient,
            secret_resolver=fake_secret_resolver,
        )
        download_candidate_products(
            actor=self.actor,
            store=self.store,
            client_factory=FakeProductsClient,
            secret_resolver=fake_secret_resolver,
        )

        basis = build_latest_product_source_basis(store=self.store, action_id="elastic-collision")

        self.assertEqual(basis["rows_count"], 3)
        self.assertEqual(basis["collision_count"], 1)
        collision = next(row for row in basis["rows"] if row["product_id"] == "401")
        self.assertEqual(collision["source_group"], "candidate_and_active")
        self.assertTrue(collision["source_details"]["collision"])
        self.assertEqual(
            collision["source_details"]["collision_reason"],
            "product_present_in_active_and_candidate_sources",
        )

    def test_permission_gating_for_task_021_product_downloads(self):
        self._select_action("elastic-permissions")
        denied = User.objects.create_user(
            login=f"observer-{self._testMethodName}",
            password="pass",
            display_name="Observer",
        )
        StoreAccess.objects.create(user=denied, store=self.store, access_level=StoreAccess.AccessLevel.VIEW)
        FakeProductsClient.active_responses = [{"result": {"products": [], "total": 0}}]

        with self.assertRaises(PermissionDenied):
            download_active_products(
                actor=denied,
                store=self.store,
                client_factory=FakeProductsClient,
                secret_resolver=fake_secret_resolver,
            )

    def _download_sources_for_product_data(self):
        self._select_action("elastic-product-data")
        FakeProductsClient.calls = []
        FakeProductsClient.active_responses = [
            {
                "result": {
                    "products": [
                        {"id": "501", "offer_id": "A-501", "min_action_price": "90", "max_action_price": "150"},
                        {"id": "502", "offer_id": "A-502", "min_action_price": "100", "max_action_price": "160"},
                        {"id": "503", "offer_id": "A-503", "min_action_price": "110", "max_action_price": "170"},
                    ],
                    "total": 3,
                }
            }
        ]
        FakeProductsClient.candidate_responses = [
            {
                "result": {
                    "products": [
                        {"id": "501", "offer_id": "C-501", "min_action_price": "90", "max_action_price": "150"},
                        {"id": "504", "offer_id": "C-504", "min_action_price": "120", "max_action_price": "180"},
                    ],
                    "total": 2,
                }
            }
        ]
        download_active_products(
            actor=self.actor,
            store=self.store,
            client_factory=FakeProductsClient,
            secret_resolver=fake_secret_resolver,
        )
        download_candidate_products(
            actor=self.actor,
            store=self.store,
            client_factory=FakeProductsClient,
            secret_resolver=fake_secret_resolver,
        )

    def test_product_data_join_maps_min_price_stocks_diagnostics_and_collision(self):
        self._download_sources_for_product_data()
        FakeProductDataClient.calls = []
        FakeProductDataClient.product_info_responses = [
            {
                "result": {
                    "items": [
                        {"id": "501", "offer_id": "O-501", "name": "Product 501", "min_price": "100.50"},
                        {"id": "502", "offer_id": "O-502", "name": "Product 502"},
                        {"id": "503", "offer_id": "O-503", "name": "Product 503", "min_price": "not-a-number"},
                    ]
                }
            }
        ]
        FakeProductDataClient.stock_responses = [
            {
                "result": {
                    "items": [
                        {
                            "product_id": "501",
                            "stocks": [
                                {"type": "fbo", "present": "2", "reserved": "99"},
                                {"type": "fbs", "present": "3", "reserved": "1"},
                            ],
                        },
                        {"product_id": "502", "stocks": [{"type": "fbo", "present": "0", "reserved": "10"}]},
                        {"product_id": "503", "stocks": [{"type": "fbs", "present": "-1", "reserved": "0"}]},
                    ]
                }
            }
        ]

        operation = download_product_data(
            actor=self.actor,
            store=self.store,
            client_factory=FakeProductDataClient,
            secret_resolver=fake_secret_resolver,
        )

        self.assertEqual(operation.step_code, OperationStepCode.OZON_API_ELASTIC_PRODUCT_DATA_DOWNLOAD)
        self.assertEqual(operation.summary["action_id"], "elastic-product-data")
        self.assertEqual(operation.summary["product_count"], 4)
        self.assertEqual(operation.summary["stock_rows_count"], 4)
        self.assertEqual(operation.summary["read_page_size"], 100)
        self.assertEqual(operation.summary["min_interval_ms"], 0)
        self.assertEqual(FakeProductDataClient.calls[0]["endpoint"], "/v3/product/info/list")
        self.assertEqual(FakeProductDataClient.calls[0]["limit"], 100)
        self.assertEqual(FakeProductDataClient.calls[1]["endpoint"], "/v4/product/info/stocks")
        self.assertEqual(set(FakeProductDataClient.calls[0]["product_ids"]), {"501", "502", "503", "504"})

        rows = {row["product_id"]: row for row in operation.summary["canonical_rows"]}
        self.assertEqual(rows["501"]["source_group"], "candidate_and_active")
        self.assertTrue(rows["501"]["source_details"]["collision"])
        self.assertEqual(rows["501"]["J_min_price"], "100.5")
        self.assertEqual(rows["501"]["R_stock_present"], "5")
        self.assertEqual(rows["501"]["O_price_min_elastic"], "90")
        self.assertEqual(rows["501"]["P_price_max_elastic"], "150")
        self.assertEqual(rows["501"].get("business_reason_code", ""), "")

        self.assertEqual(rows["502"]["business_reason_code"], "missing_min_price")
        self.assertIn("J", rows["502"]["missing_fields"])
        self.assertEqual(rows["503"]["business_reason_code"], "missing_min_price")
        self.assertEqual(rows["504"]["business_reason_code"], "missing_min_price")
        self.assertIn("ozon_api_missing_product_info", rows["504"]["diagnostics"])
        self.assertIn("ozon_api_missing_stock_info", rows["504"]["diagnostics"])
        self.assertNotIn("stock", rows["501"])

        detail_codes = set(operation.detail_rows.values_list("reason_code", flat=True))
        self.assertIn("missing_min_price", detail_codes)
        self.assertIn("ozon_api_missing_product_info", detail_codes)
        dumped = str(operation.summary)
        self.assertNotIn(CLIENT_ID, dumped)
        self.assertNotIn(API_KEY, dumped)
        self.assertNotIn("Client-Id", dumped)
        self.assertNotIn("Api-Key", dumped)

    def test_product_data_missing_stock_and_non_positive_present_map_to_no_stock(self):
        self._select_action("elastic-no-stock")
        FakeProductsClient.active_responses = [
            {
                "result": {
                    "products": [
                        {"id": "601", "min_action_price": "90", "max_action_price": "150"},
                        {"id": "602", "min_action_price": "90", "max_action_price": "150"},
                    ],
                    "total": 2,
                }
            }
        ]
        FakeProductsClient.candidate_responses = [{"result": {"products": [], "total": 0}}]
        download_active_products(
            actor=self.actor,
            store=self.store,
            client_factory=FakeProductsClient,
            secret_resolver=fake_secret_resolver,
        )
        download_candidate_products(
            actor=self.actor,
            store=self.store,
            client_factory=FakeProductsClient,
            secret_resolver=fake_secret_resolver,
        )
        FakeProductDataClient.product_info_responses = [
            {
                "result": {
                    "items": [
                        {"id": "601", "min_price": "100"},
                        {"id": "602", "min_price": "100"},
                    ]
                }
            }
        ]
        FakeProductDataClient.stock_responses = [
            {"result": {"items": [{"product_id": "601", "stocks": [{"present": "0", "reserved": "0"}]}]}}
        ]

        operation = download_product_data(
            actor=self.actor,
            store=self.store,
            client_factory=FakeProductDataClient,
            secret_resolver=fake_secret_resolver,
        )

        rows = {row["product_id"]: row for row in operation.summary["canonical_rows"]}
        self.assertEqual(rows["601"]["business_reason_code"], "no_stock")
        self.assertEqual(rows["602"]["business_reason_code"], "no_stock")
        self.assertIn("ozon_api_missing_stock_info", rows["602"]["diagnostics"])
        detail_codes = list(operation.detail_rows.order_by("row_no").values_list("reason_code", flat=True))
        self.assertEqual(detail_codes, ["no_stock", "ozon_api_missing_stock_info"])

    def test_product_data_blocks_when_task_021_source_basis_is_incomplete(self):
        self._select_action("elastic-incomplete-basis")
        FakeProductsClient.active_responses = [
            {
                "result": {
                    "products": [{"id": "701", "min_action_price": "90", "max_action_price": "150"}],
                    "total": 1,
                }
            }
        ]
        download_active_products(
            actor=self.actor,
            store=self.store,
            client_factory=FakeProductsClient,
            secret_resolver=fake_secret_resolver,
        )
        FakeProductDataClient.calls = []

        with self.assertRaises(ValidationError):
            download_product_data(
                actor=self.actor,
                store=self.store,
                client_factory=FakeProductDataClient,
                secret_resolver=fake_secret_resolver,
            )

        operation = self.store.operations.order_by("-id").first()
        self.assertEqual(operation.step_code, OperationStepCode.OZON_API_ELASTIC_PRODUCT_DATA_DOWNLOAD)
        self.assertEqual(operation.summary["result_code"], "ozon_api_response_invalid")
        self.assertEqual(operation.summary["missing_groups"], ["candidate"])
        self.assertEqual(FakeProductDataClient.calls, [])

    def test_product_data_permission_gating(self):
        self._download_sources_for_product_data()
        denied = User.objects.create_user(
            login=f"no-product-data-{self._testMethodName}",
            password="pass",
            display_name="Denied",
        )
        StoreAccess.objects.create(user=denied, store=self.store, access_level=StoreAccess.AccessLevel.VIEW)

        with self.assertRaises(PermissionDenied):
            download_product_data(
                actor=denied,
                store=self.store,
                client_factory=FakeProductDataClient,
                secret_resolver=fake_secret_resolver,
            )

    def _download_sources_for_calculation(self):
        self._select_action("elastic-calculation")
        FakeProductsClient.calls = []
        FakeProductsClient.active_responses = [
            {
                "result": {
                    "products": [
                        {"id": "801", "min_action_price": "90", "max_action_price": "150", "action_price": "120"},
                        {"id": "803", "min_action_price": "90", "max_action_price": "150", "action_price": "120"},
                        {"id": "804", "min_action_price": "90", "max_action_price": "150", "action_price": "120"},
                        {"id": "806", "min_action_price": "80", "action_price": "120"},
                        {"id": "808", "min_action_price": "90", "max_action_price": "150", "action_price": "120"},
                        {"id": "809", "min_action_price": "90", "max_action_price": "150", "action_price": "120"},
                    ],
                    "total": 6,
                }
            }
        ]
        FakeProductsClient.candidate_responses = [
            {
                "result": {
                    "products": [
                        {"id": "802", "min_action_price": "120", "max_action_price": "90"},
                        {"id": "805"},
                        {"id": "807", "max_action_price": "90"},
                        {"id": "808", "min_action_price": "90", "max_action_price": "150"},
                        {"id": "809", "min_action_price": "90", "max_action_price": "150"},
                    ],
                    "total": 5,
                }
            }
        ]
        download_active_products(
            actor=self.actor,
            store=self.store,
            client_factory=FakeProductsClient,
            secret_resolver=fake_secret_resolver,
        )
        download_candidate_products(
            actor=self.actor,
            store=self.store,
            client_factory=FakeProductsClient,
            secret_resolver=fake_secret_resolver,
        )
        FakeProductDataClient.calls = []
        FakeProductDataClient.product_info_responses = [
            {
                "result": {
                    "items": [
                        {"id": "801", "offer_id": "O-801", "name": "Product 801", "min_price": "100"},
                        {"id": "802", "offer_id": "O-802", "name": "Product 802", "min_price": "100"},
                        {"id": "803", "offer_id": "O-803", "name": "Product 803"},
                        {"id": "804", "offer_id": "O-804", "name": "Product 804", "min_price": "100"},
                        {"id": "805", "offer_id": "O-805", "name": "Product 805", "min_price": "100"},
                        {"id": "806", "offer_id": "O-806", "name": "Product 806", "min_price": "100"},
                        {"id": "807", "offer_id": "O-807", "name": "Product 807", "min_price": "100"},
                        {"id": "808", "offer_id": "O-808", "name": "Product 808", "min_price": "100"},
                        {"id": "809", "offer_id": "O-809", "name": "Product 809", "min_price": "100"},
                    ]
                }
            }
        ]
        FakeProductDataClient.stock_responses = [
            {
                "result": {
                    "items": [
                        {"product_id": "801", "stocks": [{"present": "5"}]},
                        {"product_id": "802", "stocks": [{"present": "5"}]},
                        {"product_id": "803", "stocks": [{"present": "5"}]},
                        {"product_id": "804", "stocks": [{"present": "0"}]},
                        {"product_id": "805", "stocks": [{"present": "5"}]},
                        {"product_id": "806", "stocks": [{"present": "5"}]},
                        {"product_id": "807", "stocks": [{"present": "5"}]},
                        {"product_id": "808", "stocks": [{"present": "5"}]},
                        {"product_id": "809", "stocks": [{"present": "0"}]},
                    ]
                }
            }
        ]
        return download_product_data(
            actor=self.actor,
            store=self.store,
            client_factory=FakeProductDataClient,
            secret_resolver=fake_secret_resolver,
        )

    def test_calculation_covers_rules_groups_report_and_no_manual_file(self):
        self._download_sources_for_calculation()

        operation = calculate_elastic_result(actor=self.actor, store=self.store)

        self.assertEqual(operation.step_code, OperationStepCode.OZON_API_ELASTIC_CALCULATION)
        self.assertEqual(operation.mode, OperationMode.API)
        self.assertEqual(operation.marketplace, Marketplace.OZON)
        self.assertEqual(operation.operation_type, OperationType.NOT_APPLICABLE)
        self.assertEqual(operation.summary["review_state"], "not_reviewed")
        self.assertEqual(operation.summary["deactivate_confirmation_status"], "pending")
        self.assertFalse(operation.summary["manual_upload_file_created"])
        self.assertEqual(
            operation.summary["groups_count"],
            {
                "add_to_action": 1,
                "update_action_price": 2,
                "deactivate_from_action": 4,
                "skip_candidate": 2,
                "blocked": 0,
            },
        )
        rows = {row["product_id"]: row for row in operation.summary["calculation_rows"]}
        self.assertEqual(operation.summary["action_name"], "Эластичный бустинг апрель")
        self.assertEqual(operation.summary["basis"]["selected_action"]["action_name"], "Эластичный бустинг апрель")
        self.assertEqual(rows["801"]["reason_code"], "use_max_boost_price")
        self.assertEqual(rows["801"]["action_name"], "Эластичный бустинг апрель")
        self.assertEqual(rows["801"]["planned_action"], "update_action_price")
        self.assertEqual(rows["802"]["reason_code"], "use_min_price")
        self.assertEqual(rows["802"]["planned_action"], "add_to_action")
        self.assertEqual(rows["802"]["calculated_action_price"], "100")
        self.assertEqual(rows["803"]["reason_code"], "missing_min_price")
        self.assertEqual(rows["804"]["reason_code"], "no_stock")
        self.assertEqual(rows["805"]["reason_code"], "no_boost_prices")
        self.assertEqual(rows["805"]["planned_action"], "skip_candidate")
        self.assertEqual(rows["806"]["reason_code"], "below_min_price_threshold")
        self.assertEqual(rows["807"]["reason_code"], "insufficient_ozon_input_data")
        self.assertEqual(rows["807"]["planned_action"], "skip_candidate")
        self.assertEqual(rows["808"]["source_group"], "candidate_and_active")
        self.assertEqual(rows["808"]["planned_action"], "update_action_price")
        self.assertEqual(rows["809"]["source_group"], "candidate_and_active")
        self.assertEqual(rows["809"]["planned_action"], "deactivate_from_action")
        self.assertEqual(rows["809"]["deactivate_reason_code"], "no_stock")
        for product_id in ["803", "804", "806", "809"]:
            self.assertEqual(rows[product_id]["planned_action"], "deactivate_from_action")
            self.assertTrue(rows[product_id]["deactivate_required"])
            self.assertTrue(rows[product_id]["deactivate_reason_code"])
            self.assertTrue(rows[product_id]["deactivate_reason"])

        self.assertEqual(operation.output_files.count(), 1)
        report_version = operation.output_files.select_related("file_version__file").get().file_version
        self.assertEqual(report_version.file.scenario, FileObject.Scenario.OZON_API_ELASTIC_RESULT_REPORT)
        self.assertFalse(
            FileObject.objects.filter(scenario="ozon_api_elastic_manual_upload_excel").exists()
        )
        with default_storage.open(report_version.storage_path, "rb") as handle:
            workbook = load_workbook(handle, read_only=True)
            sheet = workbook["Result"]
            headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
            self.assertIn("planned action", headers)
            self.assertIn("deactivate_reason_code", headers)
            first_row = [cell.value for cell in next(sheet.iter_rows(min_row=2, max_row=2))]
            values = dict(zip(headers, first_row, strict=True))
            self.assertEqual(values["action name"], "Эластичный бустинг апрель")
            self.assertEqual(values["action_id"], "elastic-calculation")
            self.assertEqual(values["planned action"], "update_action_price")
            workbook.close()
        dumped = str(operation.summary)
        self.assertNotIn(CLIENT_ID, dumped)
        self.assertNotIn(API_KEY, dumped)
        self.assertNotIn("Client-Id", dumped)
        self.assertNotIn("Api-Key", dumped)
        self.assertFalse(
            any("activate" in call["endpoint"] or "deactivate" in call["endpoint"] for call in FakeProductsClient.calls)
        )

    def test_api_vs_excel_decision_parity_for_golden_rows(self):
        cases = [
            (None, "90", "150", "5"),
            ("100", "90", "150", "0"),
            ("100", None, None, "5"),
            ("100", "90", "150", "5"),
            ("100", "120", "90", "5"),
            ("100", "80", None, "5"),
            ("100", None, "90", "5"),
        ]

        for row_no, (j_value, o_value, p_value, r_value) in enumerate(cases, start=1):
            with self.subTest(row_no=row_no):
                excel = decide_excel_row(
                    row_no=row_no,
                    min_price=parse_decimal(j_value),
                    min_boost_price=parse_decimal(o_value),
                    max_boost_price=parse_decimal(p_value),
                    stock=parse_decimal(r_value),
                )
                api = decide_ozon_row(
                    row_no=row_no,
                    min_price=parse_decimal(j_value),
                    min_boost_price=parse_decimal(o_value),
                    max_boost_price=parse_decimal(p_value),
                    stock=parse_decimal(r_value),
                )
                self.assertEqual(api.reason_code, excel.reason_code)
                self.assertEqual(api.participates, excel.participates)
                self.assertEqual(api.final_price, excel.final_price)

    def test_calculation_permission_gating(self):
        self._download_sources_for_calculation()
        local_admin_role = Role.objects.get(code=ROLE_LOCAL_ADMIN)
        local_admin = User.objects.create_user(
            login=f"local-admin-{self._testMethodName}",
            password="pass",
            display_name="Local Admin",
            primary_role=local_admin_role,
        )
        StoreAccess.objects.create(user=local_admin, store=self.store, access_level=StoreAccess.AccessLevel.ADMIN)

        with self.assertRaises(PermissionDenied):
            calculate_elastic_result(actor=local_admin, store=self.store)

    def test_accept_result_freezes_basis_and_generates_manual_excel_with_deactivate_sheet(self):
        self._download_sources_for_calculation()
        calculation = calculate_elastic_result(actor=self.actor, store=self.store)
        operation_count = Operation.objects.count()

        reviewed = accept_elastic_result(actor=self.actor, operation=calculation)

        self.assertEqual(Operation.objects.count(), operation_count)
        self.assertEqual(reviewed.summary["review_state"], "review_pending_deactivate_confirmation")
        self.assertEqual(reviewed.summary["accepted_basis_checksum"], reviewed.summary["basis"]["basis_checksum"])
        self.assertEqual(reviewed.summary["accepted_selected_action"]["action_id"], "elastic-calculation")
        self.assertEqual(
            reviewed.summary["accepted_selected_action"]["action_name"],
            "Эластичный бустинг апрель",
        )
        self.assertTrue(reviewed.summary["accepted_calculation_snapshot"]["calculation_rows"])
        self.assertFalse(is_upload_allowed_by_review(reviewed))
        manual_version_id = reviewed.summary["manual_upload_file_version_id"]
        manual_file = FileObject.objects.get(scenario=FileObject.Scenario.OZON_API_ELASTIC_MANUAL_UPLOAD_EXCEL)
        manual_version = manual_file.versions.get(pk=manual_version_id)
        self.assertIn("Stage 1-compatible", reviewed.summary["manual_upload_file_note"])
        with default_storage.open(manual_version.storage_path, "rb") as handle:
            workbook = load_workbook(handle, read_only=True)
            self.assertIn("Товары и цены", workbook.sheetnames)
            self.assertIn("Снять с акции", workbook.sheetnames)
            sheet = workbook["Товары и цены"]
            self.assertEqual(sheet["A1"].value, "manual upload Excel по Stage 1-compatible template")
            data_rows = list(sheet.iter_rows(min_row=4, values_only=True))
            self.assertTrue(data_rows)
            first = data_rows[0]
            self.assertEqual(first[10], "Да")
            self.assertIsNotNone(first[11])
            deactivate = workbook["Снять с акции"]
            deactivate_rows = list(deactivate.iter_rows(min_row=2, values_only=True))
            self.assertEqual(len(deactivate_rows), 4)
            headers = [cell.value for cell in next(deactivate.iter_rows(min_row=1, max_row=1))]
            reason_index = headers.index("deactivate_reason_code")
            self.assertTrue(all(row[reason_index] for row in deactivate_rows))
            note = workbook["Примечание"]
            self.assertIn("Stage 1-compatible", note["A1"].value)
            workbook.close()
        self.assertTrue(
            AuditRecord.objects.filter(
                action_code=AuditActionCode.OZON_API_ELASTIC_RESULT_REVIEWED,
                entity_type="CalculationResult",
                entity_id=str(reviewed.pk),
            ).exists()
        )

    def test_decline_sets_declined_blocks_upload_and_does_not_generate_manual_excel(self):
        self._download_sources_for_calculation()
        calculation = calculate_elastic_result(actor=self.actor, store=self.store)
        operation_count = Operation.objects.count()

        reviewed = decline_elastic_result(actor=self.actor, operation=calculation)

        self.assertEqual(Operation.objects.count(), operation_count)
        self.assertEqual(reviewed.summary["review_state"], "declined")
        self.assertEqual(reviewed.summary["accepted_basis_checksum"], "")
        self.assertFalse(reviewed.summary["accepted_calculation_snapshot"])
        self.assertFalse(is_upload_allowed_by_review(reviewed))
        self.assertFalse(
            FileObject.objects.filter(scenario=FileObject.Scenario.OZON_API_ELASTIC_MANUAL_UPLOAD_EXCEL).exists()
        )
        self.assertTrue(
            AuditRecord.objects.filter(
                action_code=AuditActionCode.OZON_API_ELASTIC_RESULT_REVIEWED,
                after_snapshot__review_state="declined",
            ).exists()
        )

    def test_review_requires_permission_and_object_access(self):
        self._download_sources_for_calculation()
        calculation = calculate_elastic_result(actor=self.actor, store=self.store)
        local_admin_role = Role.objects.get(code=ROLE_LOCAL_ADMIN)
        local_admin = User.objects.create_user(
            login=f"review-local-admin-{self._testMethodName}",
            password="pass",
            display_name="Review Local Admin",
            primary_role=local_admin_role,
        )
        StoreAccess.objects.create(user=local_admin, store=self.store, access_level=StoreAccess.AccessLevel.ADMIN)
        no_access = User.objects.create_user(
            login=f"review-no-access-{self._testMethodName}",
            password="pass",
            display_name="Review No Access",
            primary_role=Role.objects.get(code=ROLE_MARKETPLACE_MANAGER),
        )

        with self.assertRaises(PermissionDenied):
            accept_elastic_result(actor=local_admin, operation=calculation)
        with self.assertRaises(PermissionDenied):
            accept_elastic_result(actor=no_access, operation=calculation)

    def test_source_refresh_marks_accepted_result_stale(self):
        self._download_sources_for_calculation()
        calculation = calculate_elastic_result(actor=self.actor, store=self.store)
        reviewed = accept_elastic_result(actor=self.actor, operation=calculation)
        self.assertEqual(reviewed.summary["review_state"], "review_pending_deactivate_confirmation")

        FakeProductsClient.active_responses = [
            {
                "result": {
                    "products": [
                        {"id": "801", "min_action_price": "90", "max_action_price": "150", "action_price": "120"},
                    ],
                    "total": 1,
                }
            }
        ]
        download_active_products(
            actor=self.actor,
            store=self.store,
            client_factory=FakeProductsClient,
            secret_resolver=fake_secret_resolver,
        )

        reviewed.refresh_from_db()
        self.assertEqual(reviewed.summary["review_state"], "stale")
        self.assertFalse(is_upload_allowed_by_review(reviewed))

    def test_calculation_closed_catalog_has_no_undocumented_planned_or_reason_codes(self):
        self._download_sources_for_calculation()
        operation = calculate_elastic_result(actor=self.actor, store=self.store)

        allowed_reasons = {
            "missing_min_price",
            "no_stock",
            "no_boost_prices",
            "use_max_boost_price",
            "use_min_price",
            "below_min_price_threshold",
            "insufficient_ozon_input_data",
        }
        allowed_actions = {
            "add_to_action",
            "update_action_price",
            "deactivate_from_action",
            "skip_candidate",
            "blocked",
        }
        for row in operation.summary["calculation_rows"]:
            self.assertIn(row["reason_code"], allowed_reasons)
            self.assertIn(row["planned_action"], allowed_actions)

    def _accepted_for_upload(self):
        self._download_sources_for_calculation()
        calculation = calculate_elastic_result(actor=self.actor, store=self.store)
        return accept_elastic_result(actor=self.actor, operation=calculation)

    def _prepare_upload_client(self, *, drift=False, partial=False, batch_size=100):
        FakeUploadClient.calls = []
        FakeUploadClient.policy = OzonApiPolicy(
            max_read_retries=2,
            backoff_seconds=0,
            min_interval_seconds=0,
            write_batch_size=batch_size,
        )
        FakeUploadClient.actions_responses = []
        FakeUploadClient.active_membership_responses = []
        FakeUploadClient.candidate_membership_responses = []
        FakeUploadClient.product_info_responses = []
        FakeUploadClient.stock_responses = []
        FakeUploadClient.actions_response = {
            "result": [
                {
                    "id": "elastic-calculation",
                    "title": "Other" if drift == "action" else "Эластичный бустинг апрель",
                    "action_type": ELASTIC_ACTION_TYPE,
                }
            ]
        }
        FakeUploadClient.active_response = {
            "result": {
                "products": [
                    {"id": "801", "min_action_price": "90", "max_action_price": "150", "action_price": "120"},
                    {"id": "803", "min_action_price": "90", "max_action_price": "150", "action_price": "120"},
                    {"id": "804", "min_action_price": "90", "max_action_price": "150", "action_price": "120"},
                    {"id": "806", "min_action_price": "80", "action_price": "120"},
                    {"id": "808", "min_action_price": "90", "max_action_price": "150", "action_price": "120"},
                    {"id": "809", "min_action_price": "90", "max_action_price": "150", "action_price": "120"},
                ],
                "total": 6,
            }
        }
        FakeUploadClient.candidate_response = {
            "result": {
                "products": [
                    {"id": "802", "min_action_price": "120", "max_action_price": "90"},
                    {"id": "805"},
                    {"id": "807", "max_action_price": "90"},
                    {"id": "808", "min_action_price": "90", "max_action_price": "150"},
                    {"id": "809", "min_action_price": "90", "max_action_price": "150"},
                ],
                "total": 5,
            }
        }
        min_801 = "101" if drift == "j" else "100"
        FakeUploadClient.product_info_response = {
            "result": {
                "items": [
                    {"id": "801", "offer_id": "O-801", "name": "Product 801", "min_price": min_801},
                    {"id": "802", "offer_id": "O-802", "name": "Product 802", "min_price": "100"},
                    {"id": "803", "offer_id": "O-803", "name": "Product 803"},
                    {"id": "804", "offer_id": "O-804", "name": "Product 804", "min_price": "100"},
                    {"id": "805", "offer_id": "O-805", "name": "Product 805", "min_price": "100"},
                    {"id": "806", "offer_id": "O-806", "name": "Product 806", "min_price": "100"},
                    {"id": "807", "offer_id": "O-807", "name": "Product 807", "min_price": "100"},
                    {"id": "808", "offer_id": "O-808", "name": "Product 808", "min_price": "100"},
                    {"id": "809", "offer_id": "O-809", "name": "Product 809", "min_price": "100"},
                ]
            }
        }
        stock_801 = "6" if drift == "r" else "5"
        FakeUploadClient.stock_response = {
            "result": {
                "items": [
                    {"product_id": "801", "stocks": [{"present": stock_801}]},
                    {"product_id": "802", "stocks": [{"present": "5"}]},
                    {"product_id": "803", "stocks": [{"present": "5"}]},
                    {"product_id": "804", "stocks": [{"present": "0"}]},
                    {"product_id": "805", "stocks": [{"present": "5"}]},
                    {"product_id": "806", "stocks": [{"present": "5"}]},
                    {"product_id": "807", "stocks": [{"present": "5"}]},
                    {"product_id": "808", "stocks": [{"present": "5"}]},
                    {"product_id": "809", "stocks": [{"present": "0"}]},
                ]
            }
        }
        FakeUploadClient.activate_responses = [
            {
                "result": {
                    "rejected": ([{"product_id": "802", "reason": "safe rejected"}] if partial else [])
                }
            }
        ]
        FakeUploadClient.deactivate_responses = [{"result": {"rejected": []}}]

    def test_upload_blocked_without_accepted_result_or_add_update_confirmation(self):
        self._download_sources_for_calculation()
        calculation = calculate_elastic_result(actor=self.actor, store=self.store)

        with self.assertRaises(ValidationError):
            upload_elastic_result(
                actor=self.owner,
                operation=calculation,
                add_update_confirmed=True,
                client_factory=FakeUploadClient,
                secret_resolver=fake_secret_resolver,
            )
        self.assertFalse(
            Operation.objects.filter(step_code=OperationStepCode.OZON_API_ELASTIC_UPLOAD).exists()
        )

        accepted = self._accepted_for_upload()
        confirm_deactivate_group(actor=self.owner, operation=accepted)
        with self.assertRaises(ValidationError):
            upload_elastic_result(
                actor=self.owner,
                operation=accepted,
                add_update_confirmed=False,
                client_factory=FakeUploadClient,
                secret_resolver=fake_secret_resolver,
            )

    def test_deactivate_confirmation_absent_blocks_without_operation_or_write(self):
        accepted = self._accepted_for_upload()
        operation_count = Operation.objects.count()
        FakeUploadClient.calls = []

        with self.assertRaises(OzonElasticDeactivateUnconfirmed):
            upload_elastic_result(
                actor=self.owner,
                operation=accepted,
                add_update_confirmed=True,
                client_factory=FakeUploadClient,
                secret_resolver=fake_secret_resolver,
            )

        accepted.refresh_from_db()
        self.assertEqual(Operation.objects.count(), operation_count)
        self.assertEqual(accepted.summary["review_state"], "review_pending_deactivate_confirmation")
        self.assertEqual(
            accepted.summary["upload_blocked_reason_code"],
            "ozon_api_upload_blocked_deactivate_unconfirmed",
        )
        self.assertEqual(FakeUploadClient.calls, [])
        self.assertTrue(
            AuditRecord.objects.filter(
                action_code=AuditActionCode.OZON_API_ELASTIC_UPLOAD_BLOCKED_DEACTIVATE_UNCONFIRMED,
            ).exists()
        )

    def test_deactivate_confirmation_preview_and_group_confirm(self):
        accepted = self._accepted_for_upload()

        preview = deactivate_confirmation_preview(accepted)
        self.assertEqual(len(preview), 4)
        self.assertTrue(all(row["deactivate_reason_code"] for row in preview))
        confirmed = confirm_deactivate_group(actor=self.owner, operation=accepted)

        self.assertEqual(confirmed.summary["review_state"], "accepted")
        self.assertEqual(confirmed.summary["deactivate_confirmation_status"], "confirmed")
        self.assertEqual(len(confirmed.summary["deactivate_confirmation_rows"]), 4)
        self.assertTrue(
            AuditRecord.objects.filter(
                action_code=AuditActionCode.OZON_API_ELASTIC_DEACTIVATE_GROUP_CONFIRMED,
            ).exists()
        )

    def test_upload_drift_blocks_before_write_for_action_j_and_r(self):
        for drift in ("action", "j", "r"):
            with self.subTest(drift=drift):
                accepted = self._accepted_for_upload()
                confirm_deactivate_group(actor=self.owner, operation=accepted)
                self._prepare_upload_client(drift=drift)

                with self.assertRaises(OzonElasticDriftDetected):
                    upload_elastic_result(
                        actor=self.owner,
                        operation=accepted,
                        add_update_confirmed=True,
                        client_factory=FakeUploadClient,
                        secret_resolver=fake_secret_resolver,
                    )

                self.assertFalse(
                    any(
                        call["endpoint"] in {"/v1/actions/products/activate", "/v1/actions/products/deactivate"}
                        for call in FakeUploadClient.calls
                    )
                )
                upload_operation = Operation.objects.filter(
                    step_code=OperationStepCode.OZON_API_ELASTIC_UPLOAD,
                    summary__result_code="ozon_api_upload_blocked_by_drift",
                ).latest("id")
                self.assertEqual(upload_operation.status, "completed_with_error")

    def test_upload_drift_paginates_action_and_membership_until_relevant_rows(self):
        accepted = self._accepted_for_upload()
        confirm_deactivate_group(actor=self.owner, operation=accepted)
        self._prepare_upload_client()
        FakeUploadClient.actions_responses = [
            {
                "result": [
                    {"id": f"regular-{index}", "title": "Regular", "action_type": "OTHER"}
                    for index in range(100)
                ]
            },
            {
                "result": [
                    {
                        "id": "elastic-calculation",
                        "title": "Эластичный бустинг апрель",
                        "action_type": ELASTIC_ACTION_TYPE,
                    }
                ]
            },
        ]
        active_products = FakeUploadClient.active_response["result"]["products"]
        candidate_products = FakeUploadClient.candidate_response["result"]["products"]
        FakeUploadClient.active_membership_responses = [
            {
                "result": {
                    "products": [{"id": f"a-page1-{index}", "min_action_price": "1"} for index in range(100)],
                    "total": 106,
                }
            },
            {"result": {"products": active_products, "total": 106}},
        ]
        FakeUploadClient.candidate_membership_responses = [
            {
                "result": {
                    "products": [{"id": f"c-page1-{index}", "min_action_price": "1"} for index in range(100)],
                    "total": 101,
                }
            },
            {"result": {"products": candidate_products, "total": 101}},
        ]

        upload = upload_elastic_result(
            actor=self.owner,
            operation=accepted,
            add_update_confirmed=True,
            client_factory=FakeUploadClient,
            secret_resolver=fake_secret_resolver,
        )

        self.assertEqual(upload.summary["result_code"], "ozon_api_upload_success")
        action_calls = [call for call in FakeUploadClient.calls if call["endpoint"] == "/v1/actions"]
        active_calls = [call for call in FakeUploadClient.calls if call["endpoint"] == "/v1/actions/products"]
        candidate_calls = [call for call in FakeUploadClient.calls if call["endpoint"] == "/v1/actions/candidates"]
        self.assertEqual(len(action_calls), 2)
        self.assertEqual(len(active_calls), 2)
        self.assertEqual(len(candidate_calls), 2)

    def test_upload_drift_blocks_when_current_action_row_o_or_p_changed(self):
        cases = (
            ("active_o", "active_response", "801", "min_action_price", "91"),
            ("candidate_p", "candidate_response", "802", "max_action_price", "91"),
        )
        for case, response_attr, product_id, field_name, value in cases:
            with self.subTest(case=case):
                accepted = self._accepted_for_upload()
                confirm_deactivate_group(actor=self.owner, operation=accepted)
                self._prepare_upload_client()
                response = getattr(FakeUploadClient, response_attr)
                for product in response["result"]["products"]:
                    if str(product["id"]) == product_id:
                        product[field_name] = value

                with self.assertRaises(OzonElasticDriftDetected):
                    upload_elastic_result(
                        actor=self.owner,
                        operation=accepted,
                        add_update_confirmed=True,
                        client_factory=FakeUploadClient,
                        secret_resolver=fake_secret_resolver,
                    )

                self.assertFalse(
                    any(
                        call["endpoint"] in {"/v1/actions/products/activate", "/v1/actions/products/deactivate"}
                        for call in FakeUploadClient.calls
                    )
                )

    def test_upload_drift_blocks_when_add_candidate_became_active(self):
        accepted = self._accepted_for_upload()
        confirm_deactivate_group(actor=self.owner, operation=accepted)
        self._prepare_upload_client()
        FakeUploadClient.candidate_response["result"]["products"] = [
            product
            for product in FakeUploadClient.candidate_response["result"]["products"]
            if str(product["id"]) != "802"
        ]
        FakeUploadClient.candidate_response["result"]["total"] = len(
            FakeUploadClient.candidate_response["result"]["products"]
        )
        FakeUploadClient.active_response["result"]["products"].append(
            {"id": "802", "min_action_price": "120", "max_action_price": "90", "action_price": "100"}
        )
        FakeUploadClient.active_response["result"]["total"] = len(FakeUploadClient.active_response["result"]["products"])

        with self.assertRaises(OzonElasticDriftDetected):
            upload_elastic_result(
                actor=self.owner,
                operation=accepted,
                add_update_confirmed=True,
                client_factory=FakeUploadClient,
                secret_resolver=fake_secret_resolver,
            )

        self.assertFalse(
            any(
                call["endpoint"] in {"/v1/actions/products/activate", "/v1/actions/products/deactivate"}
                for call in FakeUploadClient.calls
            )
        )
        upload_operation = Operation.objects.filter(
            step_code=OperationStepCode.OZON_API_ELASTIC_UPLOAD,
            summary__result_code="ozon_api_upload_blocked_by_drift",
        ).latest("id")
        self.assertEqual(upload_operation.status, "completed_with_error")

    def _product_info_rows_for_upload_snapshot(self, rows):
        info_rows = []
        stock_rows = []
        for row in rows:
            product_id = str(row["product_id"])
            info = {"id": product_id}
            if row.get("J_min_price") not in (None, ""):
                info["min_price"] = row["J_min_price"]
            info_rows.append(info)
            stock_rows.append({"product_id": product_id, "stocks": [{"present": row.get("R_stock_present") or "0"}]})
        return info_rows, stock_rows

    def test_upload_drift_fetches_product_info_and_stocks_in_read_page_chunks(self):
        accepted = self._accepted_for_upload()
        summary = accepted.summary
        rows = list(summary["accepted_calculation_snapshot"]["calculation_rows"])
        rows.extend(
            {
                "action_id": "elastic-calculation",
                "product_id": str(9000 + index),
                "source_group": "candidate",
                "J_min_price": "100",
                "O_price_min_elastic": "90",
                "P_price_max_elastic": "80",
                "R_stock_present": "5",
                "planned_action": "skip_candidate",
                "calculated_action_price": "",
            }
            for index in range(101)
        )
        summary["accepted_calculation_snapshot"]["calculation_rows"] = rows
        Operation._base_manager.filter(pk=accepted.pk).update(summary=summary)
        accepted.refresh_from_db()
        confirm_deactivate_group(actor=self.owner, operation=accepted)
        self._prepare_upload_client()
        info_rows, stock_rows = self._product_info_rows_for_upload_snapshot(rows)
        FakeUploadClient.product_info_responses = [
            {"result": {"items": info_rows[:100]}},
            {"result": {"items": info_rows[100:]}},
        ]
        FakeUploadClient.stock_responses = [
            {"result": {"items": stock_rows[:100]}},
            {"result": {"items": stock_rows[100:]}},
        ]

        upload_elastic_result(
            actor=self.owner,
            operation=accepted,
            add_update_confirmed=True,
            client_factory=FakeUploadClient,
            secret_resolver=fake_secret_resolver,
        )

        info_calls = [call for call in FakeUploadClient.calls if call["endpoint"] == "/v3/product/info/list"]
        stock_calls = [call for call in FakeUploadClient.calls if call["endpoint"] == "/v4/product/info/stocks"]
        self.assertEqual([len(call["product_ids"]) for call in info_calls], [100, 10])
        self.assertEqual([len(call["product_ids"]) for call in stock_calls], [100, 10])

    def test_confirmed_upload_maps_payload_batches_partial_and_duplicate_protection(self):
        accepted = self._accepted_for_upload()
        confirm_deactivate_group(actor=self.owner, operation=accepted)
        self._prepare_upload_client(partial=True, batch_size=100)

        upload = upload_elastic_result(
            actor=self.owner,
            operation=accepted,
            add_update_confirmed=True,
            client_factory=FakeUploadClient,
            secret_resolver=fake_secret_resolver,
        )

        self.assertEqual(upload.step_code, OperationStepCode.OZON_API_ELASTIC_UPLOAD)
        self.assertEqual(upload.summary["result_code"], "ozon_api_upload_partial_rejected")
        self.assertEqual(upload.summary["write_batch_size"], 100)
        self.assertEqual(upload.summary["min_interval_ms"], 0)
        activate_call = next(call for call in FakeUploadClient.calls if call["endpoint"].endswith("/activate"))
        deactivate_call = next(call for call in FakeUploadClient.calls if call["endpoint"].endswith("/deactivate"))
        self.assertEqual(len(activate_call["products"]), 3)
        self.assertEqual(len(deactivate_call["products"]), 4)
        self.assertEqual(
            sorted(product["product_id"] for product in activate_call["products"]),
            ["801", "802", "808"],
        )
        self.assertTrue(all("action_price" in product for product in activate_call["products"]))
        self.assertTrue(all("action_price" not in product for product in deactivate_call["products"]))
        self.assertEqual(
            sum(1 for product in activate_call["products"] if product["product_id"] == "808"),
            1,
        )
        self.assertEqual(upload.detail_rows.filter(reason_code="ozon_api_upload_rejected").count(), 1)
        self.assertEqual(upload.output_files.count(), 1)
        self.assertEqual(
            upload.output_files.select_related("file_version__file").get().file_version.file.scenario,
            FileObject.Scenario.OZON_API_ELASTIC_UPLOAD_REPORT,
        )
        dumped = str(upload.summary)
        self.assertNotIn(CLIENT_ID, dumped)
        self.assertNotIn(API_KEY, dumped)
        self.assertNotIn("Client-Id", dumped)
        self.assertNotIn("Api-Key", dumped)
        self.assertFalse(any("/v1/product/import/prices" in str(call) for call in FakeUploadClient.calls))

        self._prepare_upload_client()
        with self.assertRaises(ValidationError):
            upload_elastic_result(
                actor=self.owner,
                operation=accepted,
                add_update_confirmed=True,
                client_factory=FakeUploadClient,
                secret_resolver=fake_secret_resolver,
            )

    def test_upload_persists_sent_batch_evidence_when_later_write_fails(self):
        accepted = self._accepted_for_upload()
        confirm_deactivate_group(actor=self.owner, operation=accepted)
        self._prepare_upload_client()
        FakeUploadClient.deactivate_responses = [OzonApiTemporaryError("timeout")]

        with self.assertRaises(OzonApiTemporaryError):
            upload_elastic_result(
                actor=self.owner,
                operation=accepted,
                add_update_confirmed=True,
                client_factory=FakeUploadClient,
                secret_resolver=fake_secret_resolver,
            )

        upload = Operation.objects.filter(
            step_code=OperationStepCode.OZON_API_ELASTIC_UPLOAD,
            summary__result_code="ozon_api_timeout",
        ).latest("id")
        self.assertEqual(upload.status, "completed_with_error")
        self.assertTrue(upload.summary["partial_evidence_persisted"])
        self.assertEqual(upload.summary["sent_batches_count"], 1)
        self.assertEqual(upload.summary["success_count"], 3)
        self.assertEqual(upload.detail_rows.filter(reason_code="ozon_api_upload_success").count(), 3)
        self.assertEqual(upload.output_files.count(), 1)
        self.assertEqual(upload.summary["batches"][0]["operation_kind"], "activate")

    def test_upload_write_batch_size_is_hard_capped_at_100(self):
        accepted = self._accepted_for_upload()
        rows = [
            {
                "action_id": "elastic-calculation",
                "product_id": str(10000 + index),
                "offer_id": f"O-{index}",
                "source_group": "candidate",
                "J_min_price": "100",
                "O_price_min_elastic": "90",
                "P_price_max_elastic": "150",
                "R_stock_present": "5",
                "planned_action": "add_to_action",
                "calculated_action_price": "150",
            }
            for index in range(101)
        ]
        summary = accepted.summary
        summary["review_state"] = "accepted"
        summary["deactivate_confirmation_status"] = "not_required"
        summary["accepted_calculation_snapshot"]["calculation_rows"] = rows
        Operation._base_manager.filter(pk=accepted.pk).update(summary=summary)
        accepted.refresh_from_db()
        self._prepare_upload_client(batch_size=500)
        FakeUploadClient.active_response = {"result": {"products": [], "total": 0}}
        FakeUploadClient.candidate_response = {
            "result": {
                "products": [
                    {"id": row["product_id"], "min_action_price": "90", "max_action_price": "150"}
                    for row in rows
                ],
                "total": len(rows),
            }
        }
        info_rows, stock_rows = self._product_info_rows_for_upload_snapshot(rows)
        FakeUploadClient.product_info_responses = [
            {"result": {"items": info_rows[:100]}},
            {"result": {"items": info_rows[100:]}},
        ]
        FakeUploadClient.stock_responses = [
            {"result": {"items": stock_rows[:100]}},
            {"result": {"items": stock_rows[100:]}},
        ]
        FakeUploadClient.activate_responses = [{"result": {"rejected": []}}, {"result": {"rejected": []}}]
        FakeUploadClient.deactivate_responses = []

        upload = upload_elastic_result(
            actor=self.owner,
            operation=accepted,
            add_update_confirmed=True,
            client_factory=FakeUploadClient,
            secret_resolver=fake_secret_resolver,
        )

        activate_calls = [call for call in FakeUploadClient.calls if call["endpoint"].endswith("/activate")]
        self.assertEqual([len(call["products"]) for call in activate_calls], [100, 1])
        self.assertEqual(upload.summary["write_batch_size"], 100)

    def test_upload_permission_gating_manager_lacks_upload_by_default(self):
        accepted = self._accepted_for_upload()

        with self.assertRaises(PermissionDenied):
            upload_elastic_result(
                actor=self.actor,
                operation=accepted,
                add_update_confirmed=True,
                client_factory=FakeUploadClient,
                secret_resolver=fake_secret_resolver,
            )
