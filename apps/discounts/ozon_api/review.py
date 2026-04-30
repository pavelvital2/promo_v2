"""Review workflow for Ozon Elastic calculation results."""

from __future__ import annotations

from copy import deepcopy
from io import BytesIO
from pathlib import Path

from django.core.exceptions import PermissionDenied, ValidationError
from django.core.files.base import ContentFile
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from apps.audit.models import AuditActionCode, AuditSourceContext
from apps.audit.services import create_audit_record
from apps.discounts.wb_api.redaction import assert_no_secret_like_values
from apps.files.models import FileObject
from apps.files.services import create_file_version
from apps.identity_access.services import has_permission
from apps.operations.models import (
    Marketplace,
    Operation,
    OperationMode,
    OperationModule,
    OperationStepCode,
    OperationType,
    ProcessStatus,
)
from apps.stores.models import StoreAccount
from apps.stores.services import require_ozon_store_for_ozon_api


REVIEW_NOT_REVIEWED = "not_reviewed"
REVIEW_ACCEPTED = "accepted"
REVIEW_DECLINED = "declined"
REVIEW_STALE = "stale"
REVIEW_PENDING_DEACTIVATE_CONFIRMATION = "review_pending_deactivate_confirmation"
REVIEW_STATES = {
    REVIEW_NOT_REVIEWED,
    REVIEW_ACCEPTED,
    REVIEW_DECLINED,
    REVIEW_STALE,
    REVIEW_PENDING_DEACTIVATE_CONFIRMATION,
}
DEACTIVATE_PENDING = "pending"
DEACTIVATE_NOT_REQUIRED = "not_required"
MANUAL_UPLOAD_LOGICAL_NAME = "ozon_api_elastic_manual_upload_excel.xlsx"
MANUAL_UPLOAD_NOTE = "Stage 1-compatible manual upload artifact for Ozon Elastic Boosting."
MANUAL_UPLOAD_TEMPLATE_PATH = (
    Path(__file__).resolve().parents[3] / "resources" / "templates" / "ozon" / "products-1977747.xlsx"
)
MANUAL_UPLOAD_DATA_START_ROW = 4
MANUAL_UPLOAD_MAX_COLUMN = 56


def _assert_review_permission(actor, store: StoreAccount) -> None:
    if not has_permission(actor, "ozon.api.elastic.review", store):
        raise PermissionDenied("No permission or object access for Ozon Elastic result review.")


def _assert_calculation_operation(operation: Operation) -> None:
    if not (
        operation.marketplace == Marketplace.OZON
        and operation.module == OperationModule.OZON_API
        and operation.mode == OperationMode.API
        and operation.operation_type == OperationType.NOT_APPLICABLE
        and operation.step_code == OperationStepCode.OZON_API_ELASTIC_CALCULATION
    ):
        raise ValidationError("Ozon Elastic calculation operation is required for review.")
    if operation.status not in {ProcessStatus.COMPLETED_SUCCESS, ProcessStatus.COMPLETED_WITH_WARNINGS}:
        raise ValidationError("Only successful Ozon Elastic calculation result can be reviewed.")


def _reviewable_summary(operation: Operation) -> dict:
    summary = deepcopy(operation.summary) if isinstance(operation.summary, dict) else {}
    state = summary.get("review_state", REVIEW_NOT_REVIEWED)
    if state not in REVIEW_STATES:
        raise ValidationError("Unsupported Ozon Elastic review state.")
    return summary


def _save_review_summary(operation: Operation, summary: dict) -> Operation:
    Operation._base_manager.filter(pk=operation.pk).update(
        summary=summary,
        updated_at=timezone.now(),
    )
    operation.refresh_from_db()
    return operation


def _accepted_snapshot(summary: dict) -> dict:
    basis = deepcopy(summary.get("basis") if isinstance(summary.get("basis"), dict) else {})
    rows = deepcopy(summary.get("calculation_rows") or [])
    snapshot = {
        "action_id": str(summary.get("action_id") or basis.get("selected_action", {}).get("action_id") or ""),
        "action_name": str(summary.get("action_name") or basis.get("selected_action", {}).get("action_name") or ""),
        "selected_action": deepcopy(basis.get("selected_action") or {}),
        "basis": basis,
        "calculation_rows": rows,
        "accepted_basis_candidate": deepcopy(summary.get("accepted_basis_candidate") or {}),
        "groups_count": deepcopy(summary.get("groups_count") or {}),
    }
    assert_no_secret_like_values(snapshot, field_name="Ozon Elastic accepted calculation snapshot")
    return snapshot


def _add_update_rows(snapshot: dict) -> list[dict]:
    return [
        row
        for row in snapshot.get("calculation_rows", [])
        if row.get("planned_action") in {"add_to_action", "update_action_price"}
    ]


def _manual_upload_rows(snapshot: dict) -> list[dict]:
    return list(snapshot.get("calculation_rows") or [])


def _deactivate_rows(snapshot: dict) -> list[dict]:
    return [
        row
        for row in snapshot.get("calculation_rows", [])
        if row.get("planned_action") == "deactivate_from_action"
    ]


def _manual_upload_template_workbook():
    if not MANUAL_UPLOAD_TEMPLATE_PATH.exists():
        raise ValidationError("Ozon manual upload Excel template is not configured.")
    workbook = load_workbook(MANUAL_UPLOAD_TEMPLATE_PATH, read_only=False, data_only=False)
    if "Товары и цены" not in workbook.sheetnames:
        workbook.close()
        raise ValidationError("Ozon manual upload Excel template must contain sheet Товары и цены.")
    sheet = workbook["Товары и цены"]
    if sheet.max_row >= MANUAL_UPLOAD_DATA_START_ROW:
        sheet.delete_rows(MANUAL_UPLOAD_DATA_START_ROW, sheet.max_row - MANUAL_UPLOAD_DATA_START_ROW + 1)
    return workbook, sheet


def _set_manual_row(sheet, row_no: int, row: dict) -> None:
    # Keep the Ozon cabinet template layout: only data row cells are populated.
    sheet.cell(row=row_no, column=1, value=row.get("product_id"))
    sheet.cell(row=row_no, column=3, value=row.get("offer_id"))
    sheet.cell(row=row_no, column=5, value=row.get("name"))
    sheet.cell(row=row_no, column=9, value=row.get("current_action_price"))
    sheet.cell(row=row_no, column=10, value=row.get("J_min_price"))
    if row.get("upload_ready"):
        sheet.cell(row=row_no, column=11, value="Да")
        sheet.cell(row=row_no, column=12, value=row.get("calculated_action_price"))
    else:
        sheet.cell(row=row_no, column=11, value=None)
        sheet.cell(row=row_no, column=12, value=None)
    sheet.cell(row=row_no, column=15, value=row.get("O_price_min_elastic"))
    sheet.cell(row=row_no, column=16, value=row.get("P_price_max_elastic"))
    sheet.cell(row=row_no, column=18, value=row.get("R_stock_present"))
    for column in range(1, MANUAL_UPLOAD_MAX_COLUMN + 1):
        cell = sheet.cell(row=row_no, column=column)
        if cell.value is None:
            cell.value = ""


def _write_manual_upload_excel(*, snapshot: dict, store: StoreAccount, actor, operation: Operation):
    workbook, sheet = _manual_upload_template_workbook()
    workbook.properties.title = MANUAL_UPLOAD_NOTE
    workbook.properties.subject = "manual upload Excel по Stage 1-compatible template"
    for row_no, row in enumerate(_manual_upload_rows(snapshot), start=4):
        _set_manual_row(sheet, row_no, row)

    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    content = ContentFile(buffer.getvalue(), name=MANUAL_UPLOAD_LOGICAL_NAME)
    return create_file_version(
        store=store,
        uploaded_by=actor,
        uploaded_file=content,
        scenario=FileObject.Scenario.OZON_API_ELASTIC_MANUAL_UPLOAD_EXCEL,
        kind=FileObject.Kind.OUTPUT,
        logical_name=MANUAL_UPLOAD_LOGICAL_NAME,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        module=OperationModule.OZON_API,
        operation_ref=operation.visible_id,
        run_ref=operation.run.visible_id,
    )


def is_upload_allowed_by_review(operation: Operation) -> bool:
    summary = operation.summary if isinstance(operation.summary, dict) else {}
    return (
        summary.get("review_state") == REVIEW_ACCEPTED
        and bool(summary.get("accepted_basis_checksum"))
        and bool(summary.get("accepted_calculation_snapshot"))
    )


@transaction.atomic
def accept_elastic_result(*, actor, operation: Operation) -> Operation:
    operation = Operation.objects.select_for_update().select_related("store", "run").get(pk=operation.pk)
    _assert_calculation_operation(operation)
    require_ozon_store_for_ozon_api(operation.store)
    _assert_review_permission(actor, operation.store)
    summary = _reviewable_summary(operation)
    previous_state = summary.get("review_state", REVIEW_NOT_REVIEWED)
    if previous_state in {REVIEW_DECLINED, REVIEW_STALE}:
        raise ValidationError("Declined or stale Ozon Elastic result cannot be accepted.")
    snapshot = _accepted_snapshot(summary)
    deactivate_rows = _deactivate_rows(snapshot)
    deactivate_status = summary.get("deactivate_confirmation_status") or (
        DEACTIVATE_PENDING if deactivate_rows else DEACTIVATE_NOT_REQUIRED
    )
    review_state = (
        REVIEW_PENDING_DEACTIVATE_CONFIRMATION
        if deactivate_rows and deactivate_status != "confirmed"
        else REVIEW_ACCEPTED
    )
    manual_version = _write_manual_upload_excel(
        snapshot=snapshot,
        store=operation.store,
        actor=actor,
        operation=operation,
    )
    reviewed_at = timezone.now()
    summary.update(
        {
            "review_state": review_state,
            "reviewed_by_user_id": actor.pk,
            "reviewed_at": reviewed_at.isoformat(),
            "accepted_basis_checksum": snapshot.get("basis", {}).get("basis_checksum") or "",
            "accepted_calculation_snapshot": snapshot,
            "accepted_selected_action": {
                "action_id": snapshot.get("action_id"),
                "action_name": snapshot.get("action_name"),
            },
            "deactivate_confirmation_status": deactivate_status,
            "manual_upload_file_created": True,
            "manual_upload_file_version_id": manual_version.pk,
            "manual_upload_file_note": "manual upload Excel по Stage 1-compatible template",
        }
    )
    assert_no_secret_like_values(summary, field_name="Ozon Elastic reviewed summary")
    operation = _save_review_summary(operation, summary)
    create_audit_record(
        action_code=AuditActionCode.OZON_API_ELASTIC_RESULT_REVIEWED,
        entity_type="CalculationResult",
        entity_id=operation.pk,
        user=actor,
        store=operation.store,
        operation=operation,
        safe_message="Ozon API Elastic result accepted.",
        before_snapshot={"review_state": previous_state},
        after_snapshot={
            "review_state": review_state,
            "accepted_basis_checksum": summary["accepted_basis_checksum"],
            "action_id": snapshot.get("action_id"),
            "action_name": snapshot.get("action_name"),
            "manual_upload_file_version_id": manual_version.pk,
            "deactivate_rows_count": len(deactivate_rows),
        },
        source_context=AuditSourceContext.SERVICE,
    )
    return operation


@transaction.atomic
def decline_elastic_result(*, actor, operation: Operation) -> Operation:
    operation = Operation.objects.select_for_update().select_related("store").get(pk=operation.pk)
    _assert_calculation_operation(operation)
    require_ozon_store_for_ozon_api(operation.store)
    _assert_review_permission(actor, operation.store)
    summary = _reviewable_summary(operation)
    previous_state = summary.get("review_state", REVIEW_NOT_REVIEWED)
    if previous_state == REVIEW_STALE:
        raise ValidationError("Stale Ozon Elastic result cannot be declined.")
    reviewed_at = timezone.now()
    summary.update(
        {
            "review_state": REVIEW_DECLINED,
            "reviewed_by_user_id": actor.pk,
            "reviewed_at": reviewed_at.isoformat(),
            "accepted_basis_checksum": "",
            "accepted_calculation_snapshot": {},
            "accepted_selected_action": {},
            "upload_blocked": True,
        }
    )
    assert_no_secret_like_values(summary, field_name="Ozon Elastic declined summary")
    operation = _save_review_summary(operation, summary)
    create_audit_record(
        action_code=AuditActionCode.OZON_API_ELASTIC_RESULT_REVIEWED,
        entity_type="CalculationResult",
        entity_id=operation.pk,
        user=actor,
        store=operation.store,
        operation=operation,
        safe_message="Ozon API Elastic result declined.",
        before_snapshot={"review_state": previous_state},
        after_snapshot={"review_state": REVIEW_DECLINED, "upload_blocked": True},
        source_context=AuditSourceContext.SERVICE,
    )
    return operation


@transaction.atomic
def mark_accepted_results_stale(*, store: StoreAccount, action_id: str | None = None, actor=None) -> int:
    queryset = Operation.objects.select_for_update().filter(
        marketplace=Marketplace.OZON,
        module=OperationModule.OZON_API,
        mode=OperationMode.API,
        operation_type=OperationType.NOT_APPLICABLE,
        step_code=OperationStepCode.OZON_API_ELASTIC_CALCULATION,
        store=store,
        summary__review_state__in=[REVIEW_ACCEPTED, REVIEW_PENDING_DEACTIVATE_CONFIRMATION],
    )
    if action_id:
        queryset = queryset.filter(summary__action_id=str(action_id))
    count = 0
    for operation in queryset:
        summary = _reviewable_summary(operation)
        previous_state = summary.get("review_state")
        summary["review_state"] = REVIEW_STALE
        summary["stale_marked_at"] = timezone.now().isoformat()
        summary["upload_blocked"] = True
        operation = _save_review_summary(operation, summary)
        create_audit_record(
            action_code=AuditActionCode.OZON_API_ELASTIC_RESULT_REVIEWED,
            entity_type="CalculationResult",
            entity_id=operation.pk,
            user=actor,
            store=store,
            operation=operation,
            safe_message="Ozon API Elastic accepted result marked stale after source refresh.",
            before_snapshot={"review_state": previous_state},
            after_snapshot={"review_state": REVIEW_STALE, "action_id": summary.get("action_id")},
            source_context=AuditSourceContext.SERVICE,
        )
        count += 1
    return count
