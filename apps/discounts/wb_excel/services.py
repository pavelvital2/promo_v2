"""WB discounts Excel parsing, calculation and operation integration."""

from __future__ import annotations

from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import PurePosixPath
from zipfile import BadZipFile

from django.core.exceptions import ValidationError
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.db import transaction
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from apps.files.models import FileObject, FileVersion
from apps.files.services import create_file_version
from apps.operations.models import (
    Marketplace,
    MessageLevel,
    OperationDetailRow,
)
from apps.operations.listing_enrichment import enrich_detail_row_marketplace_listing
from apps.operations.services import (
    InputFileSpec,
    ParameterSnapshotSpec,
    ShellExecutionResult,
    create_check_operation,
    press_process_sync,
    run_check_sync,
)
from apps.marketplace_products.services import sync_products_for_operation
from apps.platform_settings.models import StoreParameterValue, SystemParameterValue
from apps.discounts.wb_shared.calculation import decide_wb_discount


LOGIC_VERSION = "wb_discounts_excel_v1"
MAX_PROMO_FILES = 20
MAX_FILE_SIZE = 25 * 1024 * 1024
MAX_TOTAL_SIZE = 100 * 1024 * 1024

PRICE_ROLE = "price"
PROMO_ROLE = "promo"
OUTPUT_LOGICAL_NAME = "wb_discounts_output.xlsx"

REQUIRED_PRICE_COLUMNS = ("Артикул WB", "Текущая цена", "Новая скидка")
REQUIRED_PROMO_COLUMNS = (
    "Артикул WB",
    "Плановая цена для акции",
    "Загружаемая скидка для участия в акции",
)

WB_DEFAULTS = {
    "wb_threshold_percent": Decimal("70"),
    "wb_fallback_over_threshold_percent": Decimal("55"),
    "wb_fallback_no_promo_percent": Decimal("55"),
}

ROW_ERROR_CODES = {
    "wb_missing_article",
    "wb_invalid_current_price",
    "wb_duplicate_price_article",
    "wb_missing_required_column",
    "wb_invalid_workbook",
    "wb_output_write_error",
    "wb_discount_out_of_range",
}


@dataclass(frozen=True)
class AppliedWbParameter:
    code: str
    value: Decimal
    source: str
    parameter_version: str
    effective_at: object | None = None


@dataclass(frozen=True)
class WbParameters:
    threshold_percent: Decimal
    fallback_no_promo_percent: Decimal
    fallback_over_threshold_percent: Decimal
    snapshots: tuple[AppliedWbParameter, ...]


@dataclass
class PriceRow:
    row_no: int
    article: str
    current_price: Decimal | None
    errors: list[str] = field(default_factory=list)
    problem_field: str = ""


@dataclass
class PromoAggregate:
    min_discount: Decimal
    max_plan_price: Decimal


@dataclass
class Detail:
    row_no: int
    article: str
    row_status: str
    reason_code: str
    message_level: str
    message: str
    problem_field: str = ""
    current_price: Decimal | None = None
    min_discount: Decimal | None = None
    max_plan_price: Decimal | None = None
    final_discount_pre_threshold: int | None = None
    final_discount: int | None = None

    def final_value_payload(self) -> dict:
        payload = {
            "current_price": decimal_to_json(self.current_price),
            "min_discount": decimal_to_json(self.min_discount),
            "max_plan_price": decimal_to_json(self.max_plan_price),
            "final_discount_pre_threshold": self.final_discount_pre_threshold,
            "final_discount": self.final_discount,
        }
        return {key: value for key, value in payload.items() if value is not None}


@dataclass
class CalculationResult:
    summary: dict
    details: list[Detail]
    final_discounts_by_row: dict[int, int]
    error_count: int
    warning_count: int


def decimal_to_json(value: Decimal | None) -> str | None:
    return None if value is None else format(value, "f")


def normalize_article(value) -> str:
    text = "" if value is None else str(value)
    text = text.strip()
    if text.endswith(".0"):
        text = text[:-2]
    return text


def parse_decimal(value) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value if value.is_finite() else None
    text = str(value).replace(" ", "").replace("\u00a0", "").replace(",", ".").strip()
    if not text:
        return None
    try:
        decimal = Decimal(text)
    except InvalidOperation:
        return None
    return decimal if decimal.is_finite() else None


def _load_first_sheet(version: FileVersion, *, read_only: bool = True):
    if not str(version.original_name).lower().endswith(".xlsx"):
        raise ValidationError("Only .xlsx files are supported.")
    if version.size > MAX_FILE_SIZE:
        raise ValidationError("One WB input file exceeds 25 MB.")
    try:
        with default_storage.open(version.storage_path, "rb") as handle:
            workbook = load_workbook(BytesIO(handle.read()), data_only=False, read_only=read_only)
    except (OSError, BadZipFile, InvalidFileException, ValueError) as exc:
        raise ValidationError("Workbook cannot be opened safely.") from exc
    if not workbook.sheetnames:
        raise ValidationError("Workbook has no sheets.")
    sheet = workbook[workbook.sheetnames[0]]
    if read_only and hasattr(sheet, "reset_dimensions"):
        sheet.reset_dimensions()
    return workbook, sheet


def _header_map(sheet) -> dict[str, int]:
    mapping = {}
    for cell in next(sheet.iter_rows(min_row=1, max_row=1), ()):
        if cell.value is None:
            continue
        name = str(cell.value).strip()
        if name:
            mapping[name] = cell.column
    return mapping


def _missing_required_columns(header: dict[str, int], required: tuple[str, ...]) -> list[str]:
    return [column for column in required if column not in header]


def _get_cell(row, column_no: int):
    return row[column_no - 1].value


def _read_price_rows(price_version: FileVersion) -> tuple[list[PriceRow], list[Detail], int]:
    try:
        workbook, sheet = _load_first_sheet(price_version)
    except ValidationError:
        detail = Detail(
            row_no=1,
            article="",
            row_status="error",
            reason_code="wb_invalid_workbook",
            message_level=MessageLevel.ERROR,
            message="Price workbook cannot be opened safely.",
            problem_field="price:workbook",
        )
        return [], [detail], 1
    try:
        header = _header_map(sheet)
        missing = _missing_required_columns(header, REQUIRED_PRICE_COLUMNS)
        if missing:
            detail = Detail(
                row_no=1,
                article="",
                row_status="error",
                reason_code="wb_missing_required_column",
                message_level=MessageLevel.ERROR,
                message=f"Missing required price columns: {', '.join(missing)}",
                problem_field="required_columns",
            )
            return [], [detail], 1

        rows = []
        details = []
        article_counts: dict[str, int] = {}
        for row in sheet.iter_rows(min_row=2):
            row_no = row[0].row
            article = normalize_article(_get_cell(row, header["Артикул WB"]))
            current_price = parse_decimal(_get_cell(row, header["Текущая цена"]))
            price_row = PriceRow(row_no=row_no, article=article, current_price=current_price)
            if not article:
                price_row.errors.append("wb_missing_article")
                price_row.problem_field = "Артикул WB"
            if current_price is None or current_price <= 0:
                price_row.errors.append("wb_invalid_current_price")
                price_row.problem_field = price_row.problem_field or "Текущая цена"
            if article:
                article_counts[article] = article_counts.get(article, 0) + 1
            rows.append(price_row)

        duplicates = {article for article, count in article_counts.items() if count > 1}
        for price_row in rows:
            if price_row.article in duplicates:
                price_row.errors.append("wb_duplicate_price_article")
                price_row.problem_field = price_row.problem_field or "Артикул WB"

        for price_row in rows:
            for code in price_row.errors:
                details.append(
                    Detail(
                        row_no=price_row.row_no,
                        article=price_row.article,
                        row_status="error",
                        reason_code=code,
                        message_level=MessageLevel.ERROR,
                        message=_message_for_code(code),
                        problem_field=price_row.problem_field,
                        current_price=price_row.current_price,
                    )
                )
        return rows, details, len(details)
    finally:
        workbook.close()


def _read_promo_aggregates(promo_versions: list[FileVersion]) -> tuple[dict[str, PromoAggregate], list[Detail], int, int]:
    aggregates: dict[str, PromoAggregate] = {}
    details: list[Detail] = []
    valid_file_count = 0
    invalid_promo_rows = 0
    for ordinal, version in enumerate(promo_versions, start=1):
        try:
            workbook, sheet = _load_first_sheet(version)
        except ValidationError:
            details.append(
                Detail(
                    row_no=1,
                    article="",
                    row_status="error",
                    reason_code="wb_invalid_workbook",
                    message_level=MessageLevel.ERROR,
                    message=f"Promo workbook #{ordinal} cannot be opened safely.",
                    problem_field=f"promo_{ordinal}:workbook",
                )
            )
            continue

        try:
            header = _header_map(sheet)
            missing = _missing_required_columns(header, REQUIRED_PROMO_COLUMNS)
            if missing:
                details.append(
                    Detail(
                        row_no=1,
                        article="",
                        row_status="error",
                        reason_code="wb_missing_required_column",
                        message_level=MessageLevel.ERROR,
                        message=f"Missing required promo columns: {', '.join(missing)}",
                        problem_field=f"promo_{ordinal}:required_columns",
                    )
                )
                continue

            valid_file_count += 1
            for row in sheet.iter_rows(min_row=2):
                row_no = row[0].row
                article = normalize_article(_get_cell(row, header["Артикул WB"]))
                plan_price = parse_decimal(_get_cell(row, header["Плановая цена для акции"]))
                discount = parse_decimal(
                    _get_cell(row, header["Загружаемая скидка для участия в акции"])
                )
                if not article or plan_price is None or discount is None:
                    invalid_promo_rows += 1
                    details.append(
                        Detail(
                            row_no=row_no,
                            article=article,
                            row_status="warning",
                            reason_code="wb_invalid_promo_row",
                            message_level=MessageLevel.WARNING_INFO,
                            message="Promo row is ignored because article or numeric values are invalid.",
                            problem_field=f"promo_{ordinal}:row",
                            min_discount=discount,
                            max_plan_price=plan_price,
                        )
                    )
                    continue
                current = aggregates.get(article)
                if current is None:
                    aggregates[article] = PromoAggregate(
                        min_discount=discount,
                        max_plan_price=plan_price,
                    )
                else:
                    current.min_discount = min(current.min_discount, discount)
                    current.max_plan_price = max(current.max_plan_price, plan_price)
        finally:
            workbook.close()

    if valid_file_count == 0:
        details.append(
            Detail(
                row_no=1,
                article="",
                row_status="error",
                reason_code="wb_invalid_workbook",
                message_level=MessageLevel.ERROR,
                message="All promo files are invalid.",
                problem_field="promo_files",
            )
        )
    return aggregates, details, valid_file_count, invalid_promo_rows


def _message_for_code(code: str) -> str:
    return {
        "wb_valid_calculated": "Discount was calculated from promo data.",
        "wb_no_promo_item": "Article is absent from valid promo rows; fallback discount was applied.",
        "wb_over_threshold": "Calculated discount exceeded threshold; fallback discount was applied.",
        "wb_missing_article": "Price row has no valid WB article.",
        "wb_invalid_current_price": "Current price is missing, invalid, or unsafe for calculation.",
        "wb_duplicate_price_article": "Price file contains a duplicate WB article.",
        "wb_discount_out_of_range": "Final discount is outside 0..100.",
    }.get(code, code)


def validate_input_file_set(price_version: FileVersion | None, promo_versions: list[FileVersion]) -> None:
    if price_version is None:
        raise ValidationError("WB run requires exactly one price file.")
    if not promo_versions:
        raise ValidationError("WB run requires at least one promo file.")
    if len(promo_versions) > MAX_PROMO_FILES:
        raise ValidationError("WB run allows no more than 20 promo files.")
    versions = [price_version, *promo_versions]
    for version in versions:
        if not str(version.original_name).lower().endswith(".xlsx"):
            raise ValidationError("WB input files must be .xlsx.")
        if version.size > MAX_FILE_SIZE:
            raise ValidationError("One WB input file exceeds 25 MB.")
    if sum(version.size for version in versions) > MAX_TOTAL_SIZE:
        raise ValidationError("WB input files exceed 100 MB in total.")


def build_input_specs(
    price_version: FileVersion | None,
    promo_versions: list[FileVersion],
) -> list[InputFileSpec]:
    specs = []
    if price_version is not None:
        specs.append(InputFileSpec(file_version=price_version, role_in_operation=PRICE_ROLE, ordinal_no=1))
    specs.extend(
        InputFileSpec(file_version=version, role_in_operation=PROMO_ROLE, ordinal_no=index)
        for index, version in enumerate(promo_versions, start=1)
    )
    return specs


def _input_composition_details(
    price_version: FileVersion | None,
    promo_versions: list[FileVersion],
) -> list[Detail]:
    details: list[Detail] = []
    if price_version is None:
        details.append(
            Detail(
                row_no=1,
                article="",
                row_status="error",
                reason_code="wb_invalid_workbook",
                message_level=MessageLevel.ERROR,
                message="WB run requires exactly one price file.",
                problem_field="price_file",
            )
        )
    if not promo_versions:
        details.append(
            Detail(
                row_no=1,
                article="",
                row_status="error",
                reason_code="wb_invalid_workbook",
                message_level=MessageLevel.ERROR,
                message="WB run requires at least one promo file.",
                problem_field="promo_files",
            )
        )
    if len(promo_versions) > MAX_PROMO_FILES:
        details.append(
            Detail(
                row_no=1,
                article="",
                row_status="error",
                reason_code="wb_invalid_workbook",
                message_level=MessageLevel.ERROR,
                message="WB run allows no more than 20 promo files.",
                problem_field="promo_files",
            )
        )

    versions = [version for version in [price_version, *promo_versions] if version is not None]
    for ordinal, version in enumerate(versions, start=1):
        role = "price" if version == price_version else "promo"
        if not str(version.original_name).lower().endswith(".xlsx"):
            details.append(
                Detail(
                    row_no=1,
                    article="",
                    row_status="error",
                    reason_code="wb_invalid_workbook",
                    message_level=MessageLevel.ERROR,
                    message="WB input files must be .xlsx.",
                    problem_field=f"{role}_{ordinal}:extension",
                )
            )
        if version.size > MAX_FILE_SIZE:
            details.append(
                Detail(
                    row_no=1,
                    article="",
                    row_status="error",
                    reason_code="wb_invalid_workbook",
                    message_level=MessageLevel.ERROR,
                    message="One WB input file exceeds 25 MB.",
                    problem_field=f"{role}_{ordinal}:size",
                )
            )
    if sum(version.size for version in versions) > MAX_TOTAL_SIZE:
        details.append(
            Detail(
                row_no=1,
                article="",
                row_status="error",
                reason_code="wb_invalid_workbook",
                message_level=MessageLevel.ERROR,
                message="WB input files exceed 100 MB in total.",
                problem_field="input_files:size_total",
            )
        )
    return details


def resolve_wb_parameters(store) -> WbParameters:
    resolved: dict[str, AppliedWbParameter] = {}
    now = None
    for code, default in WB_DEFAULTS.items():
        store_value = (
            StoreParameterValue.objects.filter(store=store, parameter_code=code, is_active=True)
            .order_by("-active_from", "-id")
            .first()
        )
        if store_value is not None:
            resolved[code] = AppliedWbParameter(
                code=code,
                value=_parameter_decimal(store_value.value, code),
                source="store",
                parameter_version=f"store:{store_value.pk}",
                effective_at=store_value.active_from,
            )
            continue
        system_value = (
            SystemParameterValue.objects.filter(parameter_code=code)
            .order_by("-active_from", "-id")
            .first()
        )
        if system_value is not None:
            resolved[code] = AppliedWbParameter(
                code=code,
                value=_parameter_decimal(system_value.value, code),
                source="system",
                parameter_version=f"system:{system_value.pk}",
                effective_at=system_value.active_from,
            )
            continue
        resolved[code] = AppliedWbParameter(
            code=code,
            value=default,
            source="system",
            parameter_version=f"default:{decimal_to_json(default)}",
            effective_at=now,
        )

    return WbParameters(
        threshold_percent=resolved["wb_threshold_percent"].value,
        fallback_no_promo_percent=resolved["wb_fallback_no_promo_percent"].value,
        fallback_over_threshold_percent=resolved["wb_fallback_over_threshold_percent"].value,
        snapshots=tuple(resolved[code] for code in sorted(resolved)),
    )


def _parameter_decimal(value, code: str) -> Decimal:
    decimal = parse_decimal(value)
    if decimal is None:
        raise ValidationError(f"WB parameter {code} has invalid numeric value.")
    return decimal


def build_parameter_specs(parameters: WbParameters) -> list[ParameterSnapshotSpec]:
    return [
        ParameterSnapshotSpec(
            parameter_code=snapshot.code,
            applied_value=decimal_to_json(snapshot.value),
            source=snapshot.source,
            parameter_version=snapshot.parameter_version,
            effective_at=snapshot.effective_at,
        )
        for snapshot in parameters.snapshots
    ]


def calculate(
    price_version: FileVersion | None,
    promo_versions: list[FileVersion],
    parameters: WbParameters,
) -> CalculationResult:
    composition_details = _input_composition_details(price_version, promo_versions)
    if composition_details:
        error_count = sum(1 for detail in composition_details if detail.message_level == MessageLevel.ERROR)
        warning_count = sum(
            1
            for detail in composition_details
            if detail.message_level in {MessageLevel.WARNING_INFO, MessageLevel.WARNING_CONFIRMABLE}
        )
        return CalculationResult(
            summary={
                "price_rows": 0,
                "promo_files": len(promo_versions),
                "valid_promo_files": 0,
                "invalid_promo_rows": 0,
                "calculated_rows": 0,
                "error_count": error_count,
                "warning_count": warning_count,
                "logic_version": LOGIC_VERSION,
                "parameters": {
                    snapshot.code: {
                        "value": decimal_to_json(snapshot.value),
                        "source": snapshot.source,
                        "parameter_version": snapshot.parameter_version,
                    }
                    for snapshot in parameters.snapshots
                },
            },
            details=composition_details,
            final_discounts_by_row={},
            error_count=error_count,
            warning_count=warning_count,
        )

    price_rows, price_details, price_error_count = _read_price_rows(price_version)
    aggregates, promo_details, valid_promo_files, invalid_promo_rows = _read_promo_aggregates(promo_versions)
    details = [*price_details, *promo_details]
    final_discounts_by_row: dict[int, int] = {}

    if valid_promo_files > 0:
        for price_row in price_rows:
            if price_row.errors:
                continue
            aggregate = aggregates.get(price_row.article)
            if aggregate is None:
                decision = decide_wb_discount(
                    current_price=price_row.current_price,
                    min_discount=None,
                    max_plan_price=None,
                    threshold_percent=parameters.threshold_percent,
                    fallback_no_promo_percent=parameters.fallback_no_promo_percent,
                    fallback_over_threshold_percent=parameters.fallback_over_threshold_percent,
                )
                min_discount = max_plan_price = None
            else:
                decision = decide_wb_discount(
                    current_price=price_row.current_price,
                    min_discount=aggregate.min_discount,
                    max_plan_price=aggregate.max_plan_price,
                    threshold_percent=parameters.threshold_percent,
                    fallback_no_promo_percent=parameters.fallback_no_promo_percent,
                    fallback_over_threshold_percent=parameters.fallback_over_threshold_percent,
                )
                min_discount = aggregate.min_discount
                max_plan_price = aggregate.max_plan_price
            final_discount = decision.final_discount
            code = decision.reason_code
            pre_threshold = decision.final_discount_pre_threshold

            row_status = "ok"
            message_level = MessageLevel.INFO
            if final_discount < 0 or final_discount > 100:
                row_status = "error"
                message_level = MessageLevel.ERROR
                code = "wb_discount_out_of_range"
            else:
                final_discounts_by_row[price_row.row_no] = final_discount

            details.append(
                Detail(
                    row_no=price_row.row_no,
                    article=price_row.article,
                    row_status=row_status,
                    reason_code=code,
                    message_level=message_level,
                    message=_message_for_code(code),
                    problem_field="Новая скидка" if code == "wb_discount_out_of_range" else "",
                    current_price=price_row.current_price,
                    min_discount=min_discount,
                    max_plan_price=max_plan_price,
                    final_discount_pre_threshold=pre_threshold,
                    final_discount=final_discount,
                )
            )

    error_count = sum(1 for detail in details if detail.message_level == MessageLevel.ERROR)
    warning_count = sum(
        1
        for detail in details
        if detail.message_level in {MessageLevel.WARNING_INFO, MessageLevel.WARNING_CONFIRMABLE}
    )
    summary = {
        "price_rows": len(price_rows),
        "promo_files": len(promo_versions),
        "valid_promo_files": valid_promo_files,
        "invalid_promo_rows": invalid_promo_rows,
        "calculated_rows": len(final_discounts_by_row),
        "error_count": error_count,
        "warning_count": warning_count,
        "logic_version": LOGIC_VERSION,
        "parameters": {
            snapshot.code: {
                "value": decimal_to_json(snapshot.value),
                "source": snapshot.source,
                "parameter_version": snapshot.parameter_version,
            }
            for snapshot in parameters.snapshots
        },
    }
    return CalculationResult(
        summary=summary,
        details=details,
        final_discounts_by_row=final_discounts_by_row,
        error_count=error_count,
        warning_count=warning_count,
    )


def _persist_details(operation, details: list[Detail]) -> None:
    for detail in details:
        OperationDetailRow.objects.create(
            operation=operation,
            row_no=detail.row_no,
            product_ref=detail.article,
            row_status=detail.row_status,
            reason_code=detail.reason_code,
            message_level=detail.message_level,
            message=detail.message,
            problem_field=detail.problem_field,
            final_value=detail.final_value_payload() or None,
        )


def _enrich_operation_detail_listings(operation) -> None:
    for row in operation.detail_rows.select_related("operation").order_by("id"):
        enrich_detail_row_marketplace_listing(row)


def _check_executor(operation) -> ShellExecutionResult:
    price_version, promo_versions = _versions_from_operation(operation)
    parameters = _parameters_from_operation(operation)
    result = calculate(price_version, promo_versions, parameters)
    _persist_details(operation, result.details)
    sync_products_for_operation(operation)
    _enrich_operation_detail_listings(operation)
    return ShellExecutionResult(
        summary=result.summary,
        error_count=result.error_count,
        warning_count=result.warning_count,
    )


def _process_executor(operation) -> ShellExecutionResult:
    price_version, promo_versions = _versions_from_operation(operation)
    parameters = _parameters_from_operation(operation)
    result = calculate(price_version, promo_versions, parameters)
    _persist_details(operation, result.details)
    sync_products_for_operation(operation)
    _enrich_operation_detail_listings(operation)
    if result.error_count:
        return ShellExecutionResult(
            summary={**result.summary, "output_created": False},
            error_count=result.error_count,
            warning_count=result.warning_count,
        )
    try:
        output_version = _write_output_workbook(
            price_version=price_version,
            final_discounts_by_row=result.final_discounts_by_row,
            store=operation.store,
            user=operation.initiator_user,
            operation_ref=operation.visible_id,
            run_ref=operation.run.visible_id,
        )
    except ValidationError:
        OperationDetailRow.objects.create(
            operation=operation,
            row_no=1,
            product_ref="",
            row_status="error",
            reason_code="wb_output_write_error",
            message_level=MessageLevel.ERROR,
            message="Process cannot safely write output workbook column.",
            problem_field="Новая скидка",
        )
        return ShellExecutionResult(
            summary={**result.summary, "output_created": False},
            error_count=result.error_count + 1,
            warning_count=result.warning_count,
        )
    return ShellExecutionResult(
        summary={**result.summary, "output_created": True},
        error_count=result.error_count,
        warning_count=result.warning_count,
        output_file_version=output_version,
    )


def _versions_from_operation(operation) -> tuple[FileVersion | None, list[FileVersion]]:
    links = operation.input_files.select_related("file_version", "file_version__file").order_by(
        "role_in_operation",
        "ordinal_no",
    )
    price_versions = [link.file_version for link in links if link.role_in_operation == PRICE_ROLE]
    promo_versions = [link.file_version for link in links if link.role_in_operation == PROMO_ROLE]
    price_version = price_versions[0] if len(price_versions) == 1 else None
    return price_version, promo_versions


def _parameters_from_operation(operation) -> WbParameters:
    snapshots = []
    by_code = {snapshot.parameter_code: snapshot for snapshot in operation.parameter_snapshots.all()}
    for code in sorted(WB_DEFAULTS):
        snapshot = by_code.get(code)
        if snapshot is None:
            raise ValidationError("Operation is missing WB parameter snapshot.")
        snapshots.append(
            AppliedWbParameter(
                code=code,
                value=_parameter_decimal(snapshot.applied_value, code),
                source=snapshot.source,
                parameter_version=snapshot.parameter_version,
                effective_at=snapshot.effective_at,
            )
        )
    values = {snapshot.code: snapshot.value for snapshot in snapshots}
    return WbParameters(
        threshold_percent=values["wb_threshold_percent"],
        fallback_no_promo_percent=values["wb_fallback_no_promo_percent"],
        fallback_over_threshold_percent=values["wb_fallback_over_threshold_percent"],
        snapshots=tuple(snapshots),
    )


def _write_output_workbook(
    *,
    price_version: FileVersion,
    final_discounts_by_row: dict[int, int],
    store,
    user,
    operation_ref: str,
    run_ref: str,
    scenario: str = FileObject.Scenario.WB_DISCOUNTS_EXCEL,
    module: str = "discounts_excel",
    logical_name: str = OUTPUT_LOGICAL_NAME,
) -> FileVersion:
    try:
        with default_storage.open(price_version.storage_path, "rb") as handle:
            workbook = load_workbook(BytesIO(handle.read()))
        sheet = workbook[workbook.sheetnames[0]]
        header = _header_map(sheet)
        missing = _missing_required_columns(header, REQUIRED_PRICE_COLUMNS)
        if missing:
            raise ValidationError("Output workbook cannot be written without required price columns.")
        target_column = header["Новая скидка"]
        for row_no, final_discount in final_discounts_by_row.items():
            sheet.cell(row=row_no, column=target_column, value=final_discount)
        buffer = BytesIO()
        workbook.save(buffer)
    except Exception as exc:
        raise ValidationError("wb_output_write_error") from exc
    finally:
        try:
            workbook.close()
        except UnboundLocalError:
            pass
    output_name = _output_name(price_version.original_name)
    content = ContentFile(buffer.getvalue(), name=output_name)
    return create_file_version(
        store=store,
        uploaded_by=user,
        uploaded_file=content,
        scenario=scenario,
        kind=FileObject.Kind.OUTPUT,
        logical_name=logical_name,
        module=module,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        operation_ref=operation_ref,
        run_ref=run_ref,
    )


def _output_name(original_name: str) -> str:
    path = PurePosixPath(original_name)
    stem = path.stem or "wb_discounts"
    return f"{stem}_processed.xlsx"


@transaction.atomic
def run_wb_check(
    *,
    store,
    initiator_user,
    price_version: FileVersion | None,
    promo_versions: list[FileVersion],
    enforce_permissions: bool = False,
):
    parameters = resolve_wb_parameters(store)
    input_specs = build_input_specs(price_version, promo_versions)
    operation = create_check_operation(
        marketplace=Marketplace.WB,
        store=store,
        initiator_user=initiator_user,
        input_files=input_specs,
        parameters=build_parameter_specs(parameters),
        logic_version=LOGIC_VERSION,
        enforce_permissions=enforce_permissions,
    )
    return run_check_sync(operation, _check_executor)


def press_wb_process(
    *,
    store,
    initiator_user,
    price_version: FileVersion | None,
    promo_versions: list[FileVersion],
    confirmed_warning_codes: list[str] | None = None,
    enforce_permissions: bool = False,
):
    parameters = resolve_wb_parameters(store)
    input_specs = build_input_specs(price_version, promo_versions)
    return press_process_sync(
        marketplace=Marketplace.WB,
        store=store,
        initiator_user=initiator_user,
        input_files=input_specs,
        parameters=build_parameter_specs(parameters),
        logic_version=LOGIC_VERSION,
        check_executor=_check_executor,
        process_executor=_process_executor,
        confirmed_warning_codes=confirmed_warning_codes,
        enforce_permissions=enforce_permissions,
    )
