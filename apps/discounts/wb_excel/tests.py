"""Tests for TASK-007 WB discounts Excel behavior."""

from __future__ import annotations

import shutil
import tempfile
from decimal import Decimal
from io import BytesIO
from zipfile import ZIP_DEFLATED, ZipFile

from django.contrib.auth import get_user_model
from django.core.exceptions import ValidationError
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.test import TestCase, override_settings
from openpyxl import Workbook, load_workbook

from apps.files.models import FileObject
from apps.files.services import create_file_version
from apps.identity_access.models import AccessEffect, StoreAccess
from apps.identity_access.seeds import ROLE_MARKETPLACE_MANAGER, seed_identity_access
from apps.operations.models import (
    CheckStatus,
    Operation,
    OperationDetailRow,
    OperationOutputFile,
    OperationType,
    ProcessStatus,
)
from apps.product_core.models import MarketplaceListing
from apps.platform_settings.models import ParameterDefinition, StoreParameterValue, SystemParameterValue
from apps.stores.models import StoreAccount

from .services import (
    LOGIC_VERSION,
    calculate,
    parse_decimal,
    press_wb_process,
    resolve_wb_parameters,
    run_wb_check,
    validate_input_file_set,
)
from .services import MAX_FILE_SIZE


class WbExcelTask007Tests(TestCase):
    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls._media_root = tempfile.mkdtemp(prefix="promo-v2-wb-tests-")
        cls.override = override_settings(MEDIA_ROOT=cls._media_root)
        cls.override.enable()

    @classmethod
    def tearDownClass(cls):
        cls.override.disable()
        shutil.rmtree(cls._media_root, ignore_errors=True)
        super().tearDownClass()

    def setUp(self):
        seed_identity_access()
        role = self._role(ROLE_MARKETPLACE_MANAGER)
        User = get_user_model()
        self.user = User.objects.create_user(
            login=f"manager-{self._testMethodName}",
            password="password",
            display_name="Manager",
            primary_role=role,
        )
        self.store = StoreAccount.objects.create(
            name=f"WB Store {self._testMethodName}",
            marketplace=StoreAccount.Marketplace.WB,
        )
        StoreAccess.objects.create(
            user=self.user,
            store=self.store,
            access_level=StoreAccess.AccessLevel.WORK,
            effect=AccessEffect.ALLOW,
        )

    @staticmethod
    def _role(code):
        from apps.identity_access.models import Role

        return Role.objects.get(code=code)

    def _xlsx_version(self, name: str, rows: list[list], *, logical_name: str = "input"):
        workbook = Workbook()
        sheet = workbook.active
        for row in rows:
            sheet.append(row)
        buffer = BytesIO()
        workbook.save(buffer)
        workbook.close()
        content = ContentFile(buffer.getvalue(), name=name)
        return create_file_version(
            store=self.store,
            uploaded_by=self.user,
            uploaded_file=content,
            scenario=FileObject.Scenario.WB_DISCOUNTS_EXCEL,
            kind=FileObject.Kind.INPUT,
            logical_name=logical_name,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def _xlsx_version_with_bad_dimension(
        self,
        name: str,
        rows: list[list],
        *,
        logical_name: str = "input",
    ):
        workbook = Workbook()
        sheet = workbook.active
        for row in rows:
            sheet.append(row)
        buffer = BytesIO()
        workbook.save(buffer)
        workbook.close()

        source = BytesIO(buffer.getvalue())
        patched = BytesIO()
        with ZipFile(source, "r") as src, ZipFile(patched, "w", ZIP_DEFLATED) as dst:
            for item in src.infolist():
                data = src.read(item.filename)
                if item.filename == "xl/worksheets/sheet1.xml":
                    data = data.replace(
                        b'<dimension ref="A1:C2"/>',
                        b'<dimension ref="A1"/>',
                        1,
                    )
                dst.writestr(item, data)
        content = ContentFile(patched.getvalue(), name=name)
        return create_file_version(
            store=self.store,
            uploaded_by=self.user,
            uploaded_file=content,
            scenario=FileObject.Scenario.WB_DISCOUNTS_EXCEL,
            kind=FileObject.Kind.INPUT,
            logical_name=logical_name,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def _raw_version(self, name: str, content: bytes, *, logical_name: str = "input"):
        return create_file_version(
            store=self.store,
            uploaded_by=self.user,
            uploaded_file=ContentFile(content, name=name),
            scenario=FileObject.Scenario.WB_DISCOUNTS_EXCEL,
            kind=FileObject.Kind.INPUT,
            logical_name=logical_name,
            content_type="application/octet-stream",
        )

    def _price(self, rows: list[list]):
        return self._xlsx_version(
            "prices.xlsx",
            [["Артикул WB", "Текущая цена", "Новая скидка", "Не менять"], *rows],
            logical_name="price",
        )

    def _promo(self, rows: list[list], name: str = "promo.xlsx"):
        return self._xlsx_version(
            name,
            [
                [
                    "Артикул WB",
                    "Плановая цена для акции",
                    "Загружаемая скидка для участия в акции",
                ],
                *rows,
            ],
            logical_name="promo",
        )

    def test_normalization_decimal_arithmetic_ceil_and_hybrid_order(self):
        price = self._price(
            [
                [" 123.0 ", "1 000,00", 0, "keep-a"],
                ["456", "1000", 0, "keep-b"],
                ["789", "100", 0, "keep-c"],
            ]
        )
        promo = self._promo(
            [
                ["123", "333,10", "80"],
                ["123", "500", "60"],
                ["456", "100", "90"],
            ]
        )

        result = calculate(price, [promo], resolve_wb_parameters(self.store))

        self.assertEqual(result.error_count, 0)
        self.assertEqual(result.final_discounts_by_row, {2: 50, 3: 55, 4: 55})
        by_row = {detail.row_no: detail for detail in result.details if detail.row_status == "ok"}
        self.assertEqual(by_row[2].reason_code, "wb_valid_calculated")
        self.assertEqual(by_row[2].final_discount_pre_threshold, 50)
        self.assertEqual(by_row[3].reason_code, "wb_over_threshold")
        self.assertEqual(by_row[4].reason_code, "wb_no_promo_item")
        self.assertIsNone(parse_decimal("not-a-number"))

    def test_read_only_parser_ignores_stale_a1_dimension_from_wb_exports(self):
        price = self._xlsx_version_with_bad_dimension(
            "prices-bad-dimension.xlsx",
            [["Артикул WB", "Текущая цена", "Новая скидка"], ["123", "1000", 0]],
            logical_name="price",
        )
        promo = self._xlsx_version_with_bad_dimension(
            "promo-bad-dimension.xlsx",
            [
                [
                    "Артикул WB",
                    "Плановая цена для акции",
                    "Загружаемая скидка для участия в акции",
                ],
                ["123", "900", "15"],
            ],
            logical_name="promo",
        )

        result = calculate(price, [promo], resolve_wb_parameters(self.store))

        self.assertEqual(result.error_count, 0)
        self.assertEqual(result.final_discounts_by_row, {2: 10})

    def test_check_writes_no_output_and_persists_closed_reason_codes(self):
        price = self._price([["123", "1000", 0, "keep"]])
        promo = self._promo([["123", "900", "15"]])

        operation = run_wb_check(
            store=self.store,
            initiator_user=self.user,
            price_version=price,
            promo_versions=[promo],
            enforce_permissions=True,
        )

        self.assertEqual(operation.status, CheckStatus.COMPLETED_NO_ERRORS)
        self.assertEqual(operation.operation_type, OperationType.CHECK)
        self.assertEqual(operation.logic_version, LOGIC_VERSION)
        self.assertEqual(OperationOutputFile.objects.count(), 0)
        self.assertFalse(hasattr(operation, "process_result"))
        self.assertEqual(operation.parameter_snapshots.count(), 3)
        reason_codes = set(operation.detail_rows.values_list("reason_code", flat=True))
        self.assertEqual(reason_codes, {"wb_valid_calculated"})
        listing = MarketplaceListing.objects.get(store=self.store, marketplace="wb", external_primary_id="123")
        detail = operation.detail_rows.get(product_ref="123")
        self.assertEqual(detail.marketplace_listing, listing)
        self.assertEqual(detail.product_ref, "123")

    def test_process_reuses_actual_check_and_writes_only_new_discount_column(self):
        price = self._price(
            [
                ["123", "1000", 1, "left-a"],
                ["456", "1000", 2, "=1+1"],
            ]
        )
        promo = self._promo([["123", "500", "60"], ["456", "100", "10"]])

        check = run_wb_check(
            store=self.store,
            initiator_user=self.user,
            price_version=price,
            promo_versions=[promo],
            enforce_permissions=True,
        )
        result = press_wb_process(
            store=self.store,
            initiator_user=self.user,
            price_version=price,
            promo_versions=[promo],
            enforce_permissions=True,
        )

        self.assertFalse(result.check_was_created)
        self.assertEqual(result.check_operation.pk, check.pk)
        self.assertEqual(result.process_operation.status, ProcessStatus.COMPLETED_SUCCESS)
        output_link = result.process_operation.output_files.get()
        with default_storage.open(output_link.file_version.storage_path, "rb") as handle:
            workbook = load_workbook(handle, data_only=False)
        sheet = workbook[workbook.sheetnames[0]]
        self.assertEqual(sheet["A2"].value, "123")
        self.assertEqual(sheet["B2"].value, "1000")
        self.assertEqual(sheet["C2"].value, 50)
        self.assertEqual(sheet["D2"].value, "left-a")
        self.assertEqual(sheet["C3"].value, 10)
        self.assertEqual(sheet["D3"].value, "=1+1")
        workbook.close()

    def test_out_of_range_is_row_error_and_blocks_process(self):
        price = self._price([["123", "100", 0, "keep"]])
        promo = self._promo([["123", "150", "10"]])

        check = run_wb_check(
            store=self.store,
            initiator_user=self.user,
            price_version=price,
            promo_versions=[promo],
        )

        self.assertEqual(check.status, CheckStatus.COMPLETED_WITH_ERRORS)
        self.assertEqual(check.error_count, 1)
        self.assertTrue(
            OperationDetailRow.objects.filter(
                operation=check,
                reason_code="wb_discount_out_of_range",
                row_status="error",
            ).exists()
        )
        with self.assertRaises(ValidationError):
            press_wb_process(
                store=self.store,
                initiator_user=self.user,
                price_version=price,
                promo_versions=[promo],
            )
        self.assertEqual(Operation.objects.filter(operation_type=OperationType.PROCESS).count(), 0)

    def test_input_file_set_requires_one_price_and_one_to_twenty_promo_files(self):
        price = self._price([["123", "100", 0, "keep"]])
        promos = [self._promo([["123", "90", "10"]], name=f"promo-{index}.xlsx") for index in range(21)]

        with self.assertRaises(ValidationError):
            validate_input_file_set(None, [promos[0]])
        with self.assertRaises(ValidationError):
            validate_input_file_set(price, [])
        with self.assertRaises(ValidationError):
            validate_input_file_set(price, promos)

    def test_business_validation_errors_complete_check_with_errors_not_interrupted(self):
        cases = []

        cases.append(("missing price and promo", None, [], {"wb_invalid_workbook"}))

        price = self._price([["123", "100", 0, "keep"]])
        promos = [self._promo([["123", "90", "10"]], name=f"too-many-{index}.xlsx") for index in range(21)]
        cases.append(("too many promo files", price, promos, {"wb_invalid_workbook"}))

        wrong_extension_price = self._raw_version("prices.txt", b"not-xlsx", logical_name="price")
        cases.append(("wrong extension", wrong_extension_price, [self._promo([["123", "90", "10"]])], {"wb_invalid_workbook"}))

        oversized_price = self._price([["123", "100", 0, "keep"]])
        oversized_price.size = MAX_FILE_SIZE + 1
        oversized_price.save(update_fields=["size"])
        cases.append(("size limit", oversized_price, [self._promo([["123", "90", "10"]])], {"wb_invalid_workbook"}))

        missing_columns_price = self._xlsx_version(
            "missing-price-columns.xlsx",
            [["Артикул WB", "Новая скидка"], ["123", 0]],
            logical_name="price",
        )
        cases.append(
            (
                "missing required price columns",
                missing_columns_price,
                [self._promo([["123", "90", "10"]])],
                {"wb_missing_required_column"},
            )
        )

        duplicate_price = self._price([["123", "100", 0, "keep"], ["123.0", "200", 0, "keep"]])
        cases.append(
            (
                "duplicate price article",
                duplicate_price,
                [self._promo([["123", "90", "10"]])],
                {"wb_duplicate_price_article"},
            )
        )

        for label, price_version, promo_versions, expected_codes in cases:
            with self.subTest(label=label):
                operation = run_wb_check(
                    store=self.store,
                    initiator_user=self.user,
                    price_version=price_version,
                    promo_versions=promo_versions,
                )
                self.assertEqual(operation.status, CheckStatus.COMPLETED_WITH_ERRORS)
                self.assertGreater(operation.error_count, 0)
                self.assertEqual(OperationOutputFile.objects.filter(operation=operation).count(), 0)
                self.assertTrue(expected_codes.issubset(set(operation.detail_rows.values_list("reason_code", flat=True))))

    def test_corrupt_xlsx_is_invalid_workbook_business_error(self):
        corrupt_price = self._raw_version("prices.xlsx", b"this is not a zip workbook", logical_name="price")
        promo = self._promo([["123", "90", "10"]])

        operation = run_wb_check(
            store=self.store,
            initiator_user=self.user,
            price_version=corrupt_price,
            promo_versions=[promo],
        )

        self.assertEqual(operation.status, CheckStatus.COMPLETED_WITH_ERRORS)
        self.assertEqual(operation.error_count, 1)
        self.assertTrue(
            operation.detail_rows.filter(
                reason_code="wb_invalid_workbook",
                problem_field="price:workbook",
            ).exists()
        )

    def test_non_finite_numeric_values_are_invalid_not_interruptions(self):
        self.assertIsNone(parse_decimal(Decimal("NaN")))
        self.assertIsNone(parse_decimal("Infinity"))
        price = self._price([["bad-price", "NaN", 0, "keep"], ["valid-no-promo", "100", 0, "keep"]])
        promo = self._promo([["bad-promo", "Infinity", "10"], ["also-bad", "90", "-Infinity"]])

        operation = run_wb_check(
            store=self.store,
            initiator_user=self.user,
            price_version=price,
            promo_versions=[promo],
        )

        self.assertEqual(operation.status, CheckStatus.COMPLETED_WITH_ERRORS)
        self.assertTrue(operation.detail_rows.filter(reason_code="wb_invalid_current_price").exists())
        self.assertEqual(operation.detail_rows.filter(reason_code="wb_invalid_promo_row").count(), 2)
        self.assertTrue(operation.detail_rows.filter(reason_code="wb_no_promo_item", product_ref="valid-no-promo").exists())

    def test_seeded_wb_system_parameter_defaults_are_used_in_snapshot(self):
        expected = {
            "wb_threshold_percent": 70,
            "wb_fallback_over_threshold_percent": 55,
            "wb_fallback_no_promo_percent": 55,
        }
        self.assertEqual(
            set(ParameterDefinition.objects.filter(code__in=expected).values_list("code", flat=True)),
            set(expected),
        )
        for code, value in expected.items():
            self.assertTrue(SystemParameterValue.objects.filter(parameter_code=code, value=value).exists())

        parameters = resolve_wb_parameters(self.store)
        snapshots = {snapshot.code: snapshot for snapshot in parameters.snapshots}
        self.assertEqual(parameters.threshold_percent, Decimal("70"))
        self.assertEqual(parameters.fallback_over_threshold_percent, Decimal("55"))
        self.assertEqual(parameters.fallback_no_promo_percent, Decimal("55"))
        for code in expected:
            self.assertEqual(snapshots[code].source, "system")
            self.assertTrue(snapshots[code].parameter_version.startswith("system:"))

    def test_store_parameter_values_override_defaults_and_change_actuality_basis(self):
        price = self._price([["123", "1000", 0, "keep"]])
        promo = self._promo([["123", "100", "90"]])

        first = run_wb_check(store=self.store, initiator_user=self.user, price_version=price, promo_versions=[promo])
        self.assertEqual(first.detail_rows.get(row_no=2).final_value["final_discount"], 55)

        StoreParameterValue.objects.create(
            store=self.store,
            parameter_code="wb_fallback_over_threshold_percent",
            value="40",
            changed_by=self.user,
        )
        second_process = press_wb_process(
            store=self.store,
            initiator_user=self.user,
            price_version=price,
            promo_versions=[promo],
        )

        self.assertTrue(second_process.check_was_created)
        self.assertNotEqual(second_process.check_operation.pk, first.pk)
        process_detail = second_process.process_operation.detail_rows.get(row_no=2)
        self.assertEqual(process_detail.final_value["final_discount"], 40)
        snapshot = second_process.process_operation.parameter_snapshots.get(
            parameter_code="wb_fallback_over_threshold_percent"
        )
        self.assertEqual(snapshot.source, "store")
