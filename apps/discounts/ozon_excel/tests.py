"""Tests for TASK-008 Ozon discounts Excel behavior."""

from __future__ import annotations

import shutil
import tempfile
from decimal import Decimal
from io import BytesIO

from django.contrib.auth import get_user_model
from django.core.exceptions import PermissionDenied, ValidationError
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.test import TestCase, override_settings
from openpyxl import Workbook, load_workbook

from apps.files.models import FileObject
from apps.files.services import create_file_version
from apps.identity_access.models import AccessEffect, StoreAccess
from apps.identity_access.seeds import ROLE_MARKETPLACE_MANAGER, ROLE_OBSERVER, seed_identity_access
from apps.operations.models import CheckStatus, Operation, OperationOutputFile, OperationType, ProcessStatus
from apps.stores.models import StoreAccount

from .services import (
    LOGIC_VERSION,
    SHEET_NAME,
    calculate,
    decide_row,
    parse_decimal,
    press_ozon_process,
    run_ozon_check,
    validate_input_file_set,
)


class OzonExcelTask008Tests(TestCase):
    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls._media_root = tempfile.mkdtemp(prefix="promo-v2-ozon-tests-")
        cls.override = override_settings(MEDIA_ROOT=cls._media_root)
        cls.override.enable()

    @classmethod
    def tearDownClass(cls):
        cls.override.disable()
        shutil.rmtree(cls._media_root, ignore_errors=True)
        super().tearDownClass()

    def setUp(self):
        seed_identity_access()
        User = get_user_model()
        self.manager = User.objects.create_user(
            login=f"ozon-manager-{self._testMethodName}",
            password="password",
            display_name="Ozon Manager",
            primary_role=self._role(ROLE_MARKETPLACE_MANAGER),
        )
        self.observer = User.objects.create_user(
            login=f"ozon-observer-{self._testMethodName}",
            password="password",
            display_name="Ozon Observer",
            primary_role=self._role(ROLE_OBSERVER),
        )
        self.store = StoreAccount.objects.create(
            name=f"Ozon Store {self._testMethodName}",
            marketplace=StoreAccount.Marketplace.OZON,
        )
        for user in [self.manager, self.observer]:
            StoreAccess.objects.create(
                user=user,
                store=self.store,
                access_level=StoreAccess.AccessLevel.WORK,
                effect=AccessEffect.ALLOW,
            )

    @staticmethod
    def _role(code):
        from apps.identity_access.models import Role

        return Role.objects.get(code=code)

    def _xlsx_version(
        self,
        name: str,
        data_rows: list[dict],
        *,
        sheet_name: str = SHEET_NAME,
        max_column: int = 18,
        logical_name: str = "input",
    ):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
        for row_no in range(1, 4):
            for column_no in range(1, max_column + 1):
                sheet.cell(row=row_no, column=column_no, value=f"service-{row_no}-{column_no}")
        for index, data in enumerate(data_rows, start=4):
            for column_no in range(1, max_column + 1):
                sheet.cell(row=index, column=column_no, value=data.get(column_no, f"keep-{index}-{column_no}"))
        buffer = BytesIO()
        workbook.save(buffer)
        workbook.close()
        return create_file_version(
            store=self.store,
            uploaded_by=self.manager,
            uploaded_file=ContentFile(buffer.getvalue(), name=name),
            scenario=FileObject.Scenario.OZON_DISCOUNTS_EXCEL,
            kind=FileObject.Kind.INPUT,
            logical_name=logical_name,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def _raw_version(self, name: str, content: bytes):
        return create_file_version(
            store=self.store,
            uploaded_by=self.manager,
            uploaded_file=ContentFile(content, name=name),
            scenario=FileObject.Scenario.OZON_DISCOUNTS_EXCEL,
            kind=FileObject.Kind.INPUT,
            logical_name="input",
            content_type="application/octet-stream",
        )

    def _valid_workbook_with_rule_rows(self):
        return self._xlsx_version(
            "ozon.xlsx",
            [
                {10: None, 11: "old-k", 12: 999, 15: 10, 16: 20, 18: 5},
                {10: 100, 11: "old-k", 12: 999, 15: 10, 16: 20, 18: 0},
                {10: 100, 11: "old-k", 12: 999, 15: None, 16: None, 18: 5},
                {10: 100, 11: "old-k", 12: 999, 15: 90, 16: 120, 18: 5},
                {10: 100, 11: "old-k", 12: 999, 15: 110, 16: 90, 18: 5},
                {10: 100, 11: "old-k", 12: 999, 15: 90, 16: 80, 18: 5},
                {10: 100, 11: "old-k", 12: 999, 15: 110, 16: None, 18: 5},
            ],
        )

    def test_decision_rules_exact_order_and_normalization(self):
        version = self._valid_workbook_with_rule_rows()

        result = calculate([version])

        self.assertEqual(result.error_count, 0)
        self.assertEqual(result.summary["data_rows"], 7)
        self.assertEqual(result.summary["participating_rows"], 2)
        reason_by_row = {detail.row_no: detail.reason_code for detail in result.details}
        self.assertEqual(
            reason_by_row,
            {
                4: "missing_min_price",
                5: "no_stock",
                6: "no_boost_prices",
                7: "use_max_boost_price",
                8: "use_min_price",
                9: "below_min_price_threshold",
                10: "insufficient_ozon_input_data",
            },
        )
        self.assertIsNone(parse_decimal("not numeric"))
        self.assertIsNone(parse_decimal("NaN"))
        self.assertEqual(parse_decimal("1 234,50"), Decimal("1234.50"))

    def test_rule_order_prefers_max_boost_before_min_boost_threshold(self):
        decision = decide_row(
            row_no=4,
            min_price=Decimal("100"),
            min_boost_price=Decimal("90"),
            max_boost_price=Decimal("120"),
            stock=Decimal("1"),
        )

        self.assertEqual(decision.reason_code, "use_max_boost_price")
        self.assertEqual(decision.final_price, Decimal("120"))

    def test_check_writes_no_output_and_stores_no_parameters(self):
        version = self._xlsx_version("ozon.xlsx", [{10: "100", 15: "90", 16: "120", 18: "5"}])

        operation = run_ozon_check(
            store=self.store,
            initiator_user=self.manager,
            input_versions=[version],
            enforce_permissions=True,
        )

        self.assertEqual(operation.status, CheckStatus.COMPLETED_NO_ERRORS)
        self.assertEqual(operation.operation_type, OperationType.CHECK)
        self.assertEqual(operation.logic_version, LOGIC_VERSION)
        self.assertEqual(operation.parameter_snapshots.count(), 0)
        self.assertEqual(OperationOutputFile.objects.count(), 0)
        self.assertEqual(operation.detail_rows.get(row_no=4).reason_code, "use_max_boost_price")

        with default_storage.open(version.storage_path, "rb") as handle:
            workbook = load_workbook(handle)
        sheet = workbook[SHEET_NAME]
        self.assertEqual(sheet.cell(row=4, column=11).value, "keep-4-11")
        self.assertEqual(sheet.cell(row=4, column=12).value, "keep-4-12")
        workbook.close()

    def test_process_reuses_actual_check_and_writes_only_k_l(self):
        version = self._valid_workbook_with_rule_rows()
        check = run_ozon_check(store=self.store, initiator_user=self.manager, input_versions=[version])

        result = press_ozon_process(
            store=self.store,
            initiator_user=self.manager,
            input_versions=[version],
            enforce_permissions=True,
        )

        self.assertFalse(result.check_was_created)
        self.assertEqual(result.check_operation.pk, check.pk)
        self.assertEqual(result.process_operation.status, ProcessStatus.COMPLETED_SUCCESS)
        output_link = result.process_operation.output_files.get()
        self.assertIsNotNone(result.process_operation.process_result.output_file_version)
        with default_storage.open(version.storage_path, "rb") as handle:
            source_workbook = load_workbook(handle, data_only=False)
        with default_storage.open(output_link.file_version.storage_path, "rb") as handle:
            output_workbook = load_workbook(handle, data_only=False)
        source_sheet = source_workbook[SHEET_NAME]
        output_sheet = output_workbook[SHEET_NAME]
        for row_no in range(1, 11):
            for column_no in range(1, 19):
                if row_no >= 4 and column_no in {11, 12}:
                    continue
                self.assertEqual(
                    output_sheet.cell(row=row_no, column=column_no).value,
                    source_sheet.cell(row=row_no, column=column_no).value,
                    msg=f"Unexpected write at row {row_no}, column {column_no}",
                )
        self.assertIsNone(output_sheet["K4"].value)
        self.assertIsNone(output_sheet["L4"].value)
        self.assertEqual(output_sheet["K7"].value, "Да")
        self.assertEqual(output_sheet["L7"].value, 120)
        self.assertEqual(output_sheet["K8"].value, "Да")
        self.assertEqual(output_sheet["L8"].value, 100)
        self.assertIsNone(output_sheet["K9"].value)
        self.assertIsNone(output_sheet["L9"].value)
        source_workbook.close()
        output_workbook.close()

    def test_process_creates_new_check_when_file_version_changes(self):
        first = self._xlsx_version("ozon.xlsx", [{10: 100, 15: 90, 16: 120, 18: 5}])
        second = self._xlsx_version("ozon.xlsx", [{10: 100, 15: 90, 16: 120, 18: 5}])
        first_check = run_ozon_check(store=self.store, initiator_user=self.manager, input_versions=[first])

        result = press_ozon_process(store=self.store, initiator_user=self.manager, input_versions=[second])

        self.assertTrue(result.check_was_created)
        self.assertNotEqual(result.check_operation.pk, first_check.pk)
        self.assertEqual(result.check_operation.input_files.get().file_version_id, second.pk)
        self.assertEqual(Operation.objects.filter(operation_type=OperationType.CHECK).count(), 2)

    def test_business_validation_errors_complete_check_with_errors(self):
        cases = [
            ("missing file", [], "input_files"),
            ("too many files", [self._valid_workbook_with_rule_rows(), self._valid_workbook_with_rule_rows()], "input_files"),
            ("wrong extension", [self._raw_version("ozon.txt", b"not-xlsx")], "input_file:extension"),
            ("corrupt xlsx", [self._raw_version("ozon.xlsx", b"this is not a zip workbook")], "workbook"),
            ("missing sheet", [self._xlsx_version("ozon.xlsx", [{10: 100}], sheet_name="Wrong")], "workbook"),
            ("missing required column", [self._xlsx_version("ozon.xlsx", [{10: 100}], max_column=17)], "required_columns"),
        ]

        for label, versions, problem_field in cases:
            with self.subTest(label=label):
                operation = run_ozon_check(
                    store=self.store,
                    initiator_user=self.manager,
                    input_versions=versions,
                )
                self.assertEqual(operation.status, CheckStatus.COMPLETED_WITH_ERRORS)
                self.assertGreater(operation.error_count, 0)
                self.assertEqual(OperationOutputFile.objects.filter(operation=operation).count(), 0)
                self.assertTrue(operation.detail_rows.filter(problem_field=problem_field).exists())

    def test_process_is_blocked_when_auto_check_has_errors(self):
        corrupt = self._raw_version("ozon.xlsx", b"not a workbook")

        with self.assertRaises(ValidationError):
            press_ozon_process(
                store=self.store,
                initiator_user=self.manager,
                input_versions=[corrupt],
            )

        self.assertEqual(Operation.objects.filter(operation_type=OperationType.PROCESS).count(), 0)
        self.assertEqual(Operation.objects.filter(operation_type=OperationType.CHECK).count(), 1)

    def test_input_file_set_requires_exactly_one_xlsx(self):
        version = self._valid_workbook_with_rule_rows()
        with self.assertRaises(ValidationError):
            validate_input_file_set([])
        with self.assertRaises(ValidationError):
            validate_input_file_set([version, version])
        with self.assertRaises(ValidationError):
            validate_input_file_set([self._raw_version("ozon.csv", b"not xlsx")])
        validate_input_file_set([version])

    def test_permissions_for_check_and_process(self):
        version = self._valid_workbook_with_rule_rows()

        with self.assertRaises(PermissionDenied):
            run_ozon_check(
                store=self.store,
                initiator_user=self.observer,
                input_versions=[version],
                enforce_permissions=True,
            )
        with self.assertRaises(PermissionDenied):
            press_ozon_process(
                store=self.store,
                initiator_user=self.observer,
                input_versions=[version],
                enforce_permissions=True,
            )
