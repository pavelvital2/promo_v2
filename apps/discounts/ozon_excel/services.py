"""Ozon discounts Excel parsing, calculation and operation integration."""

from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal
from io import BytesIO
from pathlib import PurePosixPath
from zipfile import BadZipFile

from django.core.exceptions import ValidationError
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.db import transaction
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException

from apps.files.models import FileObject, FileVersion
from apps.files.services import create_file_version
from apps.discounts.ozon_shared.calculation import (
    OzonRowDecision,
    decimal_to_json,
    decide_ozon_row,
    message_for_code,
    parse_decimal,
    problem_field_for_decision,
)
from apps.operations.models import Marketplace, MessageLevel, OperationDetailRow
from apps.operations.listing_enrichment import enrich_detail_row_marketplace_listing
from apps.operations.services import (
    InputFileSpec,
    ShellExecutionResult,
    create_check_operation,
    press_process_sync,
    run_check_sync,
)
from apps.marketplace_products.services import sync_products_for_operation


LOGIC_VERSION = "ozon_discounts_excel_v1"
MAX_FILE_SIZE = 25 * 1024 * 1024

INPUT_ROLE = "input"
OUTPUT_LOGICAL_NAME = "ozon_discounts_output.xlsx"
SHEET_NAME = "Товары и цены"

COL_MIN_PRICE = 10  # J
COL_PARTICIPATION = 11  # K
COL_FINAL_PRICE = 12  # L
COL_MIN_BOOST = 15  # O
COL_MAX_BOOST = 16  # P
COL_STOCK = 18  # R
REQUIRED_COLUMNS = (
    COL_MIN_PRICE,
    COL_PARTICIPATION,
    COL_FINAL_PRICE,
    COL_MIN_BOOST,
    COL_MAX_BOOST,
    COL_STOCK,
)


RowDecision = OzonRowDecision


@dataclass(frozen=True)
class Detail:
    row_no: int
    row_status: str
    reason_code: str
    message_level: str
    message: str
    product_ref: str = ""
    problem_field: str = ""
    final_value: dict | None = None


@dataclass(frozen=True)
class CalculationResult:
    summary: dict
    details: list[Detail]
    decisions_by_row: dict[int, RowDecision]
    error_count: int
    warning_count: int


def _message_for_code(code: str) -> str:
    return message_for_code(code)


def _critical_detail(message: str, problem_field: str) -> Detail:
    return Detail(
        row_no=1,
        row_status="error",
        reason_code="",
        message_level=MessageLevel.ERROR,
        message=message,
        problem_field=problem_field,
    )


def _load_workbook(version: FileVersion, *, read_only: bool):
    if not str(version.original_name).lower().endswith(".xlsx"):
        raise ValidationError("Only .xlsx files are supported.")
    if version.size > MAX_FILE_SIZE:
        raise ValidationError("One Ozon input file exceeds 25 MB.")
    try:
        with default_storage.open(version.storage_path, "rb") as handle:
            workbook = load_workbook(BytesIO(handle.read()), data_only=False, read_only=read_only)
    except (OSError, BadZipFile, InvalidFileException, ValueError) as exc:
        raise ValidationError("Workbook cannot be opened safely.") from exc
    if SHEET_NAME not in workbook.sheetnames:
        workbook.close()
        raise ValidationError(f"Workbook must contain sheet {SHEET_NAME}.")
    sheet = workbook[SHEET_NAME]
    if read_only and hasattr(sheet, "reset_dimensions"):
        sheet.reset_dimensions()
    return workbook, sheet


def _missing_required_columns(sheet) -> list[str]:
    if (sheet.max_column or 0) < max(REQUIRED_COLUMNS) and hasattr(sheet, "calculate_dimension"):
        sheet.calculate_dimension(force=True)
    max_column = sheet.max_column or 0
    return [get_column_letter(column_no) for column_no in REQUIRED_COLUMNS if max_column < column_no]


def validate_input_file_set(input_versions: list[FileVersion]) -> None:
    if len(input_versions) != 1:
        raise ValidationError("Ozon run requires exactly one input file.")
    version = input_versions[0]
    if not str(version.original_name).lower().endswith(".xlsx"):
        raise ValidationError("Ozon input file must be .xlsx.")
    if version.size > MAX_FILE_SIZE:
        raise ValidationError("One Ozon input file exceeds 25 MB.")


def build_input_specs(input_versions: list[FileVersion]) -> list[InputFileSpec]:
    return [
        InputFileSpec(file_version=version, role_in_operation=INPUT_ROLE, ordinal_no=index)
        for index, version in enumerate(input_versions, start=1)
    ]


def _input_composition_details(input_versions: list[FileVersion]) -> list[Detail]:
    details: list[Detail] = []
    if len(input_versions) != 1:
        details.append(_critical_detail("Ozon run requires exactly one input file.", "input_files"))
        return details
    version = input_versions[0]
    if not str(version.original_name).lower().endswith(".xlsx"):
        details.append(_critical_detail("Ozon input file must be .xlsx.", "input_file:extension"))
    if version.size > MAX_FILE_SIZE:
        details.append(_critical_detail("One Ozon input file exceeds 25 MB.", "input_file:size"))
    return details


def decide_row(
    *,
    row_no: int,
    min_price: Decimal | None,
    min_boost_price: Decimal | None,
    max_boost_price: Decimal | None,
    stock: Decimal | None,
) -> RowDecision:
    return decide_ozon_row(
        row_no=row_no,
        min_price=min_price,
        min_boost_price=min_boost_price,
        max_boost_price=max_boost_price,
        stock=stock,
    )


def calculate(input_versions: list[FileVersion]) -> CalculationResult:
    composition_details = _input_composition_details(input_versions)
    if composition_details:
        return _result_from_error_details(composition_details, input_file_count=len(input_versions))

    version = input_versions[0]
    try:
        workbook, sheet = _load_workbook(version, read_only=True)
    except ValidationError as exc:
        details = [_critical_detail(str(exc.message if hasattr(exc, "message") else exc), "workbook")]
        return _result_from_error_details(details, input_file_count=1)

    try:
        missing = _missing_required_columns(sheet)
        if missing:
            details = [
                _critical_detail(
                    f"Missing required Ozon columns: {', '.join(missing)}.",
                    "required_columns",
                )
            ]
            return _result_from_error_details(details, input_file_count=1)

        details: list[Detail] = []
        decisions_by_row: dict[int, RowDecision] = {}
        for row in sheet.iter_rows(min_row=4):
            row_no = row[0].row
            decision = decide_row(
                row_no=row_no,
                min_price=parse_decimal(row[COL_MIN_PRICE - 1].value),
                min_boost_price=parse_decimal(row[COL_MIN_BOOST - 1].value),
                max_boost_price=parse_decimal(row[COL_MAX_BOOST - 1].value),
                stock=parse_decimal(row[COL_STOCK - 1].value),
            )
            decisions_by_row[row_no] = decision
            details.append(
                Detail(
                    row_no=row_no,
                    row_status="ok",
                    reason_code=decision.reason_code,
                    message_level=MessageLevel.INFO,
                    message=_message_for_code(decision.reason_code),
                    problem_field=_problem_field_for_decision(decision),
                    final_value=decision.final_value_payload(),
                )
            )
    finally:
        workbook.close()

    participating_rows = sum(1 for decision in decisions_by_row.values() if decision.participates)
    summary = {
        "input_files": 1,
        "data_rows": len(decisions_by_row),
        "participating_rows": participating_rows,
        "skipped_rows": len(decisions_by_row) - participating_rows,
        "error_count": 0,
        "warning_count": 0,
        "logic_version": LOGIC_VERSION,
        "parameters": {},
    }
    return CalculationResult(
        summary=summary,
        details=details,
        decisions_by_row=decisions_by_row,
        error_count=0,
        warning_count=0,
    )


def _result_from_error_details(details: list[Detail], *, input_file_count: int) -> CalculationResult:
    error_count = sum(1 for detail in details if detail.message_level == MessageLevel.ERROR)
    warning_count = sum(
        1
        for detail in details
        if detail.message_level in {MessageLevel.WARNING_INFO, MessageLevel.WARNING_CONFIRMABLE}
    )
    return CalculationResult(
        summary={
            "input_files": input_file_count,
            "data_rows": 0,
            "participating_rows": 0,
            "skipped_rows": 0,
            "error_count": error_count,
            "warning_count": warning_count,
            "logic_version": LOGIC_VERSION,
            "parameters": {},
        },
        details=details,
        decisions_by_row={},
        error_count=error_count,
        warning_count=warning_count,
    )


def _problem_field_for_decision(decision: RowDecision) -> str:
    return problem_field_for_decision(decision)


def _persist_details(operation, details: list[Detail]) -> None:
    for detail in details:
        OperationDetailRow.objects.create(
            operation=operation,
            row_no=detail.row_no,
            product_ref=detail.product_ref,
            row_status=detail.row_status,
            reason_code=detail.reason_code,
            message_level=detail.message_level,
            message=detail.message,
            problem_field=detail.problem_field,
            final_value=detail.final_value,
        )


def _enrich_operation_detail_listings(operation) -> None:
    for row in operation.detail_rows.select_related("operation").order_by("id"):
        enrich_detail_row_marketplace_listing(row)


def _versions_from_operation(operation) -> list[FileVersion]:
    links = operation.input_files.select_related("file_version", "file_version__file").order_by(
        "ordinal_no",
        "id",
    )
    return [link.file_version for link in links if link.role_in_operation == INPUT_ROLE]


def _check_executor(operation) -> ShellExecutionResult:
    result = calculate(_versions_from_operation(operation))
    _persist_details(operation, result.details)
    sync_products_for_operation(operation)
    _enrich_operation_detail_listings(operation)
    return ShellExecutionResult(
        summary=result.summary,
        error_count=result.error_count,
        warning_count=result.warning_count,
    )


def _process_executor(operation) -> ShellExecutionResult:
    input_versions = _versions_from_operation(operation)
    result = calculate(input_versions)
    _persist_details(operation, result.details)
    sync_products_for_operation(operation)
    _enrich_operation_detail_listings(operation)
    if result.error_count:
        return ShellExecutionResult(
            summary={**result.summary, "output_created": False},
            error_count=result.error_count,
            warning_count=result.warning_count,
        )
    try:
        output_version = _write_output_workbook(
            input_version=input_versions[0],
            decisions_by_row=result.decisions_by_row,
            store=operation.store,
            user=operation.initiator_user,
            operation_ref=operation.visible_id,
            run_ref=operation.run.visible_id,
        )
    except ValidationError:
        OperationDetailRow.objects.create(
            operation=operation,
            row_no=1,
            product_ref="",
            row_status="error",
            reason_code="",
            message_level=MessageLevel.ERROR,
            message="Process cannot safely write output workbook columns.",
            problem_field="K/L",
        )
        return ShellExecutionResult(
            summary={**result.summary, "output_created": False},
            error_count=result.error_count + 1,
            warning_count=result.warning_count,
        )
    return ShellExecutionResult(
        summary={**result.summary, "output_created": True},
        error_count=result.error_count,
        warning_count=result.warning_count,
        output_file_version=output_version,
    )


def _write_output_workbook(
    *,
    input_version: FileVersion,
    decisions_by_row: dict[int, RowDecision],
    store,
    user,
    operation_ref: str,
    run_ref: str,
) -> FileVersion:
    try:
        workbook, sheet = _load_workbook(input_version, read_only=False)
        missing = _missing_required_columns(sheet)
        if missing:
            raise ValidationError("Output workbook cannot be written without required Ozon columns.")
        for row_no, decision in decisions_by_row.items():
            sheet.cell(row=row_no, column=COL_PARTICIPATION).value = (
                "Да" if decision.participates else None
            )
            sheet.cell(row=row_no, column=COL_FINAL_PRICE).value = decision.final_price
        buffer = BytesIO()
        workbook.save(buffer)
    except Exception as exc:
        raise ValidationError("ozon_output_write_error") from exc
    finally:
        try:
            workbook.close()
        except UnboundLocalError:
            pass

    output_name = _output_name(input_version.original_name)
    content = ContentFile(buffer.getvalue(), name=output_name)
    return create_file_version(
        store=store,
        uploaded_by=user,
        uploaded_file=content,
        scenario=FileObject.Scenario.OZON_DISCOUNTS_EXCEL,
        kind=FileObject.Kind.OUTPUT,
        logical_name=OUTPUT_LOGICAL_NAME,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        operation_ref=operation_ref,
        run_ref=run_ref,
    )


def _output_name(original_name: str) -> str:
    path = PurePosixPath(original_name)
    stem = path.stem or "ozon_discounts"
    return f"{stem}_processed.xlsx"


@transaction.atomic
def run_ozon_check(
    *,
    store,
    initiator_user,
    input_versions: list[FileVersion],
    enforce_permissions: bool = False,
):
    input_specs = build_input_specs(input_versions)
    operation = create_check_operation(
        marketplace=Marketplace.OZON,
        store=store,
        initiator_user=initiator_user,
        input_files=input_specs,
        parameters=[],
        logic_version=LOGIC_VERSION,
        enforce_permissions=enforce_permissions,
    )
    return run_check_sync(operation, _check_executor)


def press_ozon_process(
    *,
    store,
    initiator_user,
    input_versions: list[FileVersion],
    confirmed_warning_codes: list[str] | None = None,
    enforce_permissions: bool = False,
):
    input_specs = build_input_specs(input_versions)
    return press_process_sync(
        marketplace=Marketplace.OZON,
        store=store,
        initiator_user=initiator_user,
        input_files=input_specs,
        parameters=[],
        logic_version=LOGIC_VERSION,
        check_executor=_check_executor,
        process_executor=_process_executor,
        confirmed_warning_codes=confirmed_warning_codes,
        enforce_permissions=enforce_permissions,
    )
